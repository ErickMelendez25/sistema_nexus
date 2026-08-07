# ==========================================================
# acciones_post_login.py (versión FINAL ULTRA-RÁPIDA ⚡)
# ==========================================================
import os
import time
import pandas as pd
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException

# 🔹 IMPORTAMOS LA FUNCIÓN DE CARACTERÍSTICAS
try:
    from agregar_caracteristicas_auto import agregar_caracteristicas
except ImportError:
    agregar_caracteristicas = None

import os
import sys



def ruta_recurso(relativa):
    try:
        base = sys._MEIPASS  # cuando es .exe
    except:
        base = os.path.abspath(".")  # cuando es .py
    return os.path.join(base, relativa)






# ==========================================================
# 🔹 CERRAR MODALES DESPUÉS DEL LOGIN
# ==========================================================
def cerrar_ventanas_emergentes(driver):
    """Cierra ventanas emergentes tras el login con ejecución inmediata."""
    try:
        botones = driver.find_elements(By.ID, "btnCerrarMod")
        for btn in botones:
            try:
                driver.execute_script("arguments[0].click();", btn)
                print("✅ Ventana emergente cerrada rápidamente.")
            except Exception:
                pass
    except Exception:
        pass



# ==========================================================
# 🔹 NAVEGAR A PROFORMAS
# ==========================================================
def navegar_a_proformas(driver, wait):
    """Después de cerrar modales, abre el menú Proformas y Ordenes y luego ingresa a Proformas."""
    try:
        # 1️⃣ Click en "Proformas y Ordenes" (menú desplegable)
        menu_proformas = wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, "//a[contains(.,'Proformas y Ordenes') and contains(@href,'#')]")
            )
        )
        driver.execute_script("arguments[0].click();", menu_proformas)
        print("✅ Click en 'Proformas y Ordenes' realizado.")

        # 2️⃣ Click en "Proformas" dentro del submenú
        enlace_proformas = wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, "//a[@href='/t_Proforma' and contains(.,'Proformas')]")
            )
        )
        driver.execute_script("arguments[0].click();", enlace_proformas)
        print("✅ Ingreso al módulo 'Proformas' completado.")

    except Exception as e:
        print(f"⚠️ Error al navegar a Proformas: {e}")


# ==========================================================
# 🔹 SELECCIONAR COMBOS Y REALIZAR BÚSQUEDA
# ==========================================================
# ==========================================================
# 🔹 SELECCIONAR COMBOS Y REALIZAR BÚSQUEDA
# ==========================================================

# === FUNCION SELLECCIONAR COMBOS EN LA PÁGINA ===
def seleccionar_combos_y_buscar(driver, wait, acuerdo_marco_seleccionado, catalogo_seleccionado, fecha_inicial, fecha_final, MARCAS_RESTRINGIDAS=None):
    print("🧩 Seleccionando valores de los combos...")


    MARCAS_RESTRINGIDAS = {
        "ECOLIMPIA","DICALI","CICLÓN LÍDER EN EL NORTE","ANAX",
        "INDUBRILL","FACIL","PARILU","PARILÚ","INSOMED",
        "BLANUX","ATOJ","DULQUI","RAYMI",
        "AIME","DELQUIMS","KUELAP","STEEL FORJA",
        "Q MASTER","ECO CLEAN","PÓMAC","ESTRELLA DEL NORTE",
        "REGGIO","DERMA PRO","TOOLS","ISIS",
        "MUCHICK","APOLO","SINCHY","CISNE",
        "DIONNE","VICRO","TACLLA","ZENCLEAN",
        "VIANFORTPRO","BELONA","BOMELSA","PROTEGE",
        "AYCER","SILGAL","BRICEL","MAXCER","GRIUPOLY",
        "SIRYCATA","CERCOR","CETOOL","JARVIS",
        "REYSER","CIRCE","LIFAL","EL PÁRAMO PERÚ",
        "THN","ITZEL"
    }


    
        # 🔎 Ajustar ZOOM para que aparezca el botón Cotizar
    try:
        driver.execute_script("document.body.style.zoom='80%';")
        print("🔍 Zoom ajustado a 80%")
        time.sleep(0.5)
    except:
        print("⚠️ No se pudo aplicar el zoom")

    # --- Seleccionar Acuerdo Marco (Select2) ---
    try:
        contenedor = wait.until(EC.element_to_be_clickable((By.ID, "select2-cboAcuerdo-container")))
        contenedor.click()
        time.sleep(0.5)

        # Buscar opción exacta y hacer scroll hasta ella
        opcion = wait.until(EC.visibility_of_element_located(
            (By.XPATH, f"//ul[@id='select2-cboAcuerdo-results']/li[contains(text(),'{acuerdo_marco_seleccionado}')]")
        ))
        driver.execute_script("arguments[0].scrollIntoView(true);", opcion)
        opcion.click()
        print("✅ Acuerdo Marco seleccionado:", acuerdo_marco_seleccionado)
        time.sleep(1)  # Esperar a que catálogo se actualice
    except Exception as e:
        print("⚠️ No se pudo seleccionar Acuerdo Marco:", e)

    # --- Seleccionar Catálogo ---
    try:
        # Esperar a que el select del catálogo tenga más de 1 opción
        catalogo_select = wait.until(EC.presence_of_element_located((By.ID, "cboCatalogo")))
        WebDriverWait(driver, 10).until(
            lambda d: len(catalogo_select.find_elements(By.TAG_NAME, "option")) > 1
        )

        # Buscar la opción exacta
        opciones = Select(catalogo_select).options
        encontrado = False
        for i, opt in enumerate(opciones):
            if catalogo_seleccionado.strip() == opt.text.strip():
                Select(catalogo_select).select_by_index(i)
                encontrado = True
                break
        if encontrado:
            print("✅ Catálogo seleccionado:", catalogo_seleccionado)
        else:
            print("⚠️ Catálogo no encontrado:", catalogo_seleccionado)
        time.sleep(0.5)
    except Exception as e:
        print("⚠️ No se pudo seleccionar catálogo:", e)

    # --- Fechas ---
    try:
        fecha_ini_elem = wait.until(EC.presence_of_element_located((By.ID, "fechaInicial")))
        fecha_ini_elem.clear()
        fecha_ini_elem.send_keys(fecha_inicial)
        fecha_fin_elem = wait.until(EC.presence_of_element_located((By.ID, "fechaFinal")))
        fecha_fin_elem.clear()
        fecha_fin_elem.send_keys(fecha_final)
        print(f"✅ Fechas seleccionadas: {fecha_inicial} - {fecha_final}")
    except Exception as e:
        print("⚠️ No se pudieron seleccionar fechas:", e)

    # --- Estado PENDIENTE ---
    try:
        estado_select = wait.until(EC.presence_of_element_located((By.ID, "cboEstado")))
        Select(estado_select).select_by_value("1")
        print("✅ Estado seleccionado: PENDIENTE")
    except Exception as e:
        print("⚠️ No se pudo seleccionar estado:", e)



    # --- Buscar ---
    try:
        btn_buscar = wait.until(EC.element_to_be_clickable((By.ID, "btnBuscar")))
        driver.execute_script("arguments[0].click();", btn_buscar)
        print("🔍 Click en 'Iniciar Búsqueda' realizado correctamente.")
    except Exception as e:
        print("⚠️ No se pudo presionar 'Iniciar Búsqueda':", e)
        
        
        

def norm(x):
    if x is None:
        return ""
    x = str(x)
    x = x.replace("–", "-").replace("—", "-")
    x = x.replace("\n", " ").replace("\r", " ")
    x = x.strip().upper()
    if x in ("NAN", "NONE"):
        return ""
    return x 

def extraer_tabla_a_excel(driver, acuerdo_seleccionado):
    import os
    import time
    import pandas as pd
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    import openpyxl
    from datetime import datetime



    
    # ==========================================================
    # 🧠 DICCIONARIO DE MARCAS (LIMPIEZA + CEREALES + BEBIDAS)
    # ==========================================================
    MARCAS_DICC = [
        # Limpieza / Industrial / Hogar
        "W&M SIRYCATA","ITZEL","D-SANIZ","PÓMAC","MILLENNIUM","KYSER",
        "SERCONSLIMP","COVENANT","C3 P","SUPREMIO","ULTRA+","RINRI",
        "RAYMI","APOLO","BELONA","ISIS","ECOKASA","PARILU","SINCHY",
        "CIRCE","DELSA","REGGIO","D SALOME","JARVIS","CICLÓN",
        "ECOLIMPIA","SILGAL","DIONNE","SUAVISSA","PILLKO","AYCER",
        "MAXCER","ANAX","PLASTIC STEEL FORJA","STEEL FORJA","DICALI",
        "AITY","+ASEO","LAMOSA CLEAN","H TOOLS","DQ DELQUIMS",
        "NOELIA","CELESTINA","JAWASS","VIANFORTPRO","BIODECOR",
        "BLANUX","MUCHICK","YOREL","PROSERLIM","ZENCLEAN",
        "GOOD CLEANER","TACLLA","EBRIEL","JAZAY","VICRO","LIFAL",
        "RAGNAR","SALPE","REY PLAST","SMARTPLAST","CERCOR","CETOOL",
        "ATOJ","BUGUI","ESTRELLA DEL NORTE","DARYAL","DULQUI",
        "GP GRIUPOLY","OSCCONTA","INDUBRILL","KUELAP","RHOMANSA",
        "INSOMED","NEW KRAL","ESE","MAXIMASS","WYPALL","JOGRANSA",
        "DIMAGSA","PROLIM","BOMELSA","ECOSOFT","MAPIALE","TEXTILES ECOKASA",
        "THN","DARYZA","SOLMATIC","DOÑA MARIA","AGLAB","ECONORACKS",
        "FABARLI","DALHI","MAGINSA","MAFA","SCOTT","LIMCOFER",
        "WHUAYRA","INVEMATT","YANNICK","INVERCOM PERU","SFOLL",
        "DAYR","RS PASSIONES TEX","NEW AYMAX","PULIZIA","PERCUS",
        "T&R CLEANER","DERMA PRO","FACIL","HOUSE LIVING","MVILLEGAS",
        "BASA","AIME","JHANSE","ELITE","ARUBBA","WIPE MASTER",
        "ARIESS","MEGACOM","YERICO","NORT COLTON","JEKMAYCAN",
        "TOOLBOX","COMPAKTO1","SCARLETT","INSTITUCIONAL SUPER",
        "ELITE PROFESSIONAL","MELISSA","GLAX","+PRO","HI LIMP",
        "SUMAC","TOILÉ","THAILER","3Q PLASTICS","FONLEA",
        "CERMAX CLEAN","3Q COTTONS","CAPACMAYO","FBK PERÚ",
        "VIRUTEX","LACTISOFT","ALESSI","MAXTIC","PRL PARILU EXPRESS",
        "MOTITA","LEONSOL","ALICAF","LYNO","ELLIA","PARACAS",
        "CISNE","MASSIA","LA FOQUITA","PROLIMSO","CHACON","F FERNELLY",
        "SIAL","GRAFI PAPEL","J&R STEEL CP","SANIT","FONCHY",
        "FAMILY DOCTOR","BIGNER","MAYA","TUINIES","SUPER",
        "WEST MICROSAFE","AYAX","KLEENEX","HANDEEL","BRILLOL'S",
        "DERQUSA","Q MASTER","SUAVE","VIBALCA","PETALO","BUBBLE",
        
        #ACCESORIOS DOMÉSTICOS
        "LLANKAQ",
        

        # Cereales / Abarrotes
        "KASQUI","KOMILON","COSTEÑO","MENESTRERO","DEL CIELO",
        "PAISANA","CERRO GRANDE","JAPRIM","MOLINO ROJO","COSTEÑITO",
        "OLLITA","FORTILIFE","PALMA REAL","DEL NORTE","PURA CAÑA",
        "HOJA REDONDA","DEL HOGAR","DOÑA RUFI",

        # Bebidas
        "VITALIA","CIELO"
    ]

    
    def detectar_marca(ficha):
        texto = ficha.upper()
        import re

        # 1️⃣ Buscar "MARCA: XXXX" pero validar contra diccionario
        match = re.search(r"MARCA\s*:\s*([A-Z0-9 +&'.-]+)", texto)
        if match:
            posible = match.group(1).strip()

            # devolver SOLO la marca válida del diccionario
            for marca in sorted(MARCAS_DICC, key=len, reverse=True):
                if posible.startswith(marca):
                    return marca

        # 2️⃣ Fallback: buscar marca directa en el texto
        for marca in sorted(MARCAS_DICC, key=len, reverse=True):
            if f" {marca} " in f" {texto} ":
                return marca

        return ""


    


    def extraer_codigo_desde_ficha(ficha):
        import re

        if not ficha:
            return ""

        # ===============================
        # 1️⃣ NORMALIZACIÓN FUERTE
        # ===============================
        texto = ficha.upper()
        texto = texto.replace("–", "-").replace("—", "-")
        texto = re.sub(r'\s+', ' ', texto).strip()

        # ===============================
        # 2️⃣ LIMPIEZA DE BASURA SEGURA
        # ===============================
        # Quitar EAN / CUBSO largos
        texto = re.sub(r'\b\d{13,15}\b', ' ', texto)

        # Quitar rangos y porcentajes (7.50 - 8.50, 20 - 33.00)
        texto = re.sub(r'\b\d+(\.\d+)?\s*-\s*\d+(\.\d+)?\b', ' ', texto)

        # Quitar tiempos
        texto = re.sub(r'\b\d+(\.\d+)?\s*(MESES|AÑOS|AÑO)\b', ' ', texto)

        # ===============================
        # 3️⃣ TRABAJAR SOLO CON EL FINAL
        # ===============================
        corte = int(len(texto) * 0.65)
        texto_final = texto[corte:]

        # ===============================
        # 4️⃣ PATRONES ROBUSTOS (ORDEN IMPORTA)
        # ===============================
        patrones = [

            # 🔥 Con símbolos + (RECOGEDOR+FILO+369)
            r'\b[A-Z0-9]+(?:\+[A-Z0-9]+){1,}\b',

            # 🔥 Múltiples guiones (TACHO-C-1201SN)
            r'\b[A-Z0-9]+(?:-[A-Z0-9.%]+){2,}\b',

            # 🔥 Guión simple completo (RE23-A, SHP7.5-20LT, 403-4LT)
            r'\b[A-Z0-9]{2,}\s*-\s*[A-Z0-9.]+\b',

            # 🔥 Marca + número (EBRIEL 732, DARYZA 31422, 1119024 VIRUTEX)
            r'\b[A-Z]{3,}\s+\d{3,8}\b|\b\d{3,8}\s+[A-Z]{3,}\b',

            # 🔥 Alfanumérico compacto (ERPLAZ, HL20L7.5%)
            r'\b[A-Z]{2,}\d+[A-Z0-9.%]*\b',

            # 🔥 Numérico puro válido (13120043, 300107, 31423)
            r'\b\d{5,8}\b',
        ]

        # ===============================
        # 5️⃣ BUSCAR DESDE EL FINAL
        # ===============================
        for patron in patrones:
            matches = list(re.finditer(patron, texto_final))
            if matches:
                codigo = matches[-1].group(0).strip()
                return codigo

        return ""
    

    def obtener_url_imagen_desde_pdf(pdf_url):
        if not pdf_url:
            return ""

        # Cambiar carpeta
        img_url = pdf_url.replace(
            "/Documentos/Productos/",
            "/Imagenes/Productos/"
        )

        # Cambiar extensión
        img_url = img_url.rsplit(".", 1)[0] + ".jpg"

        return img_url

    

    def limpiar_valor(v):
        if v is None:
            return ""
            
        v = str(v).strip()
        
        if v.lower() in ["nan", "none"]:
            return ""
            
        return v




    def esperar_loader(driver, timeout=15):
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        
        try:
            WebDriverWait(driver, timeout).until_not(
                EC.presence_of_element_located((By.CSS_SELECTOR, ".loading, .spinner, .k-loading-mask"))
            )
            return True
        except:
            return False
        
        
            
    def esperar_loader_robusto(driver, timeout_total=40):

        inicio = time.time()

        while True:

            filas = driver.find_elements(By.CSS_SELECTOR, "#tbData .FilaDatos")

            loaders_visibles = driver.find_elements(
                By.CSS_SELECTOR,
                ".k-loading-mask:not([style*='display: none'])"
            )

            print("⏳ Loader visible:", len(loaders_visibles), "| Filas:", len(filas))

            # 🟡 CASO 1: tabla recargando
            if len(filas) == 0:
                print("⌛ Tabla recargando...")
                time.sleep(1)
                continue

            # 🟢 CASO 2: tabla lista
            if not loaders_visibles and len(filas) > 0:
                return True

            # 🧠 CASO 3: loader pegado
            if loaders_visibles and len(filas) > 0:

                print("🧠 Loader pegado detectado — eliminando loader")

                try:
                    driver.execute_script("""
                        var l = document.querySelector('.k-loading-mask');
                        if(l){ l.style.display='none'; }
                    """)
                except:
                    pass

                return True

            # 💀 CASO 4: freeze real
            if time.time() - inicio > timeout_total:
                print("💀 FREEZE REAL")
                return False

            time.sleep(0.5)
                
        
    def recuperar_de_freeze(driver, pagina_actual):

        print("♻ RECUPERANDO SISTEMA POR FREEZE...")

        try:
            driver.execute_script("window.stop();")
        except:
            pass

        try:
            driver.refresh()
            time.sleep(2)
        except:
            pass

        WebDriverWait(driver, 40).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "#tbData"))
        )

        esperar_loader_robusto(driver, 40)

        ir_a_pagina(driver, pagina_actual)

        print("✅ SISTEMA RECUPERADO")
        

        
        
        
    def cerrar_modal_forzado(driver):
        """
        Cierra el modal y espera a que la tabla principal se recargue.
        """
        try:
            boton = driver.find_element(By.ID, "btnCerraPopupCotizacion")
            if boton.is_displayed():
                driver.execute_script("arguments[0].click();", boton)
                print("🧹 Modal cerrado automáticamente")

                time.sleep(1)
                
                # 🔹 Esperar que la tabla real vuelva a estar visible
                WebDriverWait(driver, 15).until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, "#tbData .FilaDatos"))
                )
                
                # 🔹 Esperar loader desaparezca
                esperar_loader_robusto(driver, 40)
                
                return True
        except:
            pass
        return False
    
    
    def estabilizar_tabla_antes_modal(driver, pagina_actual):
        """
        Asegura que la tabla esté completamente lista
        antes de intentar abrir el modal otra vez.
        """

        # 1️⃣ Esperar loader
        if not esperar_loader_robusto(driver, 30):
            recuperar_de_freeze(driver, pagina_actual)

        # 2️⃣ Esperar que existan filas visibles
        WebDriverWait(driver, 15).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "#tbData .FilaDatos"))
        )

        # 3️⃣ Esperar que NO exista modal visible
        try:
            WebDriverWait(driver, 5).until_not(
                EC.visibility_of_element_located((By.CSS_SELECTOR, "#divItem"))
            )
        except:
            pass

        # 4️⃣ Pequeña pausa técnica (CRÍTICA)
        time.sleep(0.8)

        
        
    def refresh_inteligente(driver, pagina_objetivo):
        print("♻ Intentando recuperar sistema con refresh controlado...")

        driver.refresh()

        time.sleep(2)

        esperar_loader_robusto(driver, 30)

        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "#tbData"))
        )

        ir_a_pagina(driver, pagina_objetivo)

        print("✅ Sistema recuperado")


    # ==========================================================
    # 🔁 FUNCIÓN PARA VOLVER A UNA PÁGINA ESPECÍFICA
    # ==========================================================
    def obtener_pagina_activa(driver):
        try:
            return driver.find_element(
                By.CSS_SELECTOR,
                "#divPaginacion input.NavegaActivo"
            ).get_attribute("value")
        except:
            return "0"
        

    


    def ir_a_pagina(driver, numero_objetivo):
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        import time

        print(f"🎯 Navegando inteligentemente a página {numero_objetivo}")

        try:
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.ID, "divPaginacion"))
            )

            pagina_actual_real = int(obtener_pagina_activa(driver))

            if pagina_actual_real == numero_objetivo:
                print("🟢 Ya estamos en la página correcta")
                return True

            boton_directo = driver.find_elements(
                By.XPATH,
                f"//div[@id='divPaginacion']//input[@value='{numero_objetivo}']"
            )


            if boton_directo:

                # 🔥 ESPERAR QUE CAMBIE FIRMA DE TABLA (ULTRA ROBUSTO)

                firma_antes = obtener_firma_pagina(driver)

                driver.execute_script("arguments[0].click();", boton_directo[0])

                WebDriverWait(driver, 25).until(
                    lambda d: obtener_firma_pagina(d) != firma_antes
                )



                WebDriverWait(driver, 25).until(
                    lambda d: obtener_pagina_activa(d) == str(numero_objetivo)
                )

                # 🔥 Esperar que existan filas reales en nueva página
                WebDriverWait(driver, 25).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "#tbData .FilaDatos"))
                )

                time.sleep(1.2)

                return True

            botones_visibles = driver.find_elements(
                By.XPATH,
                "//div[@id='divPaginacion']//input[@type='button' and not(contains(@value,'Siguiente')) and not(contains(@value,'Anterior'))]"
            )

            numeros_visibles = []

            for b in botones_visibles:
                try:
                    val = int(b.get_attribute("value"))
                    numeros_visibles.append(val)
                except:
                    pass

            if not numeros_visibles:
                return False

            ancla = max([n for n in numeros_visibles if n < numero_objetivo], default=None)

            if not ancla:
                return False

            boton_ancla = driver.find_element(
                By.XPATH,
                f"//div[@id='divPaginacion']//input[@value='{ancla}']"
            )

            driver.execute_script("arguments[0].click();", boton_ancla)

            WebDriverWait(driver, 20).until(
                lambda d: obtener_pagina_activa(d) == str(ancla)
            )

            time.sleep(1)

            pagina_actual_real = ancla

            while pagina_actual_real < numero_objetivo:
                boton_siguiente = driver.find_element(
                    By.XPATH,
                    "//div[@id='divPaginacion']//input[contains(@value,'Siguiente')]"
                )

                driver.execute_script("arguments[0].click();", boton_siguiente)

                WebDriverWait(driver, 20).until(
                    lambda d: obtener_pagina_activa(d) == str(pagina_actual_real + 1)
                )

                pagina_actual_real += 1
                time.sleep(0.8)

            return True

        except Exception as e:
            print(f"💀 Error navegando a página {numero_objetivo}: {e}")
            return False


    from selenium.common.exceptions import StaleElementReferenceException, TimeoutException

    def ir_a_siguiente_pagina_robusto(driver, pagina_actual):

        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        import time

        print("\n🔎 Intentando ir a la siguiente página...")

        for intento in range(5):  # 🔥 reintentos anti-stale

            try:

                # 🧠 Esperar que paginador exista
                WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.ID, "divPaginacion"))
                )

                # 🔑 firma actual de la tabla
                firma_antes = obtener_firma_pagina(driver)

                # 🔥 RECAPTURAR botón SIEMPRE
                boton_siguiente = driver.find_element(
                    By.XPATH,
                    "//div[@id='divPaginacion']//input[contains(@value,'Siguiente')]"
                )

                # 🧠 detectar última página
                if not boton_siguiente.is_enabled():
                    print("🏁 Última página detectada")
                    return False

                # scroll por seguridad
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", boton_siguiente)

                time.sleep(0.3)

                driver.execute_script("arguments[0].click();", boton_siguiente)

                # 🔥 esperar que cambie la tabla
                WebDriverWait(driver, 25).until(
                    lambda d: obtener_firma_pagina(d) != firma_antes
                )

                print("✅ Página siguiente cargada")

                return True

            except StaleElementReferenceException:

                print(f"♻ Stale detectado — reintentando ({intento+1}/5)")
                time.sleep(1)

            except TimeoutException:

                print("⚠ Timeout esperando cambio de página")

                # verificar si ya cambió igual
                firma_despues = obtener_firma_pagina(driver)

                if firma_despues != firma_antes:
                    print("✅ Página cambió igual")
                    return True

                time.sleep(1)

            except Exception as e:

                print("⚠ Error paginando:", e)
                time.sleep(1)

        print("💀 No se pudo paginar después de varios intentos")
        return False
        
    
    
    from selenium.common.exceptions import (
    StaleElementReferenceException,
    ElementClickInterceptedException,
    TimeoutException
    )

    def click_cotizar_robusto(driver, idx, pagina_actual, max_intentos=5):

        for intento in range(max_intentos):
            try:
                # 🔄 SIEMPRE recapturar filas
                # 🛡️ ESPERAR QUE TABLA TERMINE DE CARGAR
                if not esperar_loader_robusto(driver, 35):
                    recuperar_de_freeze(driver, pagina_actual)

                filas = WebDriverWait(driver, 15).until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, "#tbData .FilaDatos"))
                )


                try:
                    fila = filas[idx]
                except IndexError:
                    print("⚠ Fila ya no existe tras recarga — continuando")
                    continue

                boton = fila.find_element(By.CSS_SELECTOR, "button[title='Cotizar']")

                WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable(boton)
                )

                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", boton)
                time.sleep(0.5)

                driver.execute_script("arguments[0].click();", boton)

                return True

            except (StaleElementReferenceException, ElementClickInterceptedException, TimeoutException):

                print(f"🔄 Reintentando click ({intento+1}/{max_intentos})")

                esperar_loader(driver, 15)
                ir_a_pagina(driver, pagina_actual)

                time.sleep(1)

        return False



    # ==========================================================
    # 🔑 FUNCIÓN FIRMA DE PÁGINA (CONTROL DE PAGINACIÓN)
    # ==========================================================
    def obtener_firma_pagina(driver):
        filas = driver.find_elements(By.CSS_SELECTOR, "#tbData tr.FilaDatos")
        if not filas:
            return None
        return filas[0].find_element(By.CSS_SELECTOR, "td.primerTd").text.strip()


    print("\n📊 INICIANDO EXTRACCIÓN…")



    try:
        driver.execute_script("document.body.style.zoom='80%'")
    except:
        pass

    ruta_excel = ruta_recurso("assets/EXTRAER.xlsx")
    
    
    columnas = [
    "N°","FECHA_GUARDADO","PAGINA", "FILA_PAGINA", "REQUERIMIENTO", "PROCEDIMIENTO", "PROFORMA",
    "FECHA DE EMISION", "FECHA LIMITE DE COTIZACION",
    "ENTIDAD", "RUC", "PRODUCTO", "FICHA PRODUCTO",
    "MARCA", "CODIGO", "CANTIDAD", "DIRECCION",
    "INICIO DE ENTREGA", "PLAZO MAXI", "FIN DE ENTREGA",
    "SUBTOTAL", "PDF", "IMAGEN" # 👈 NUEVA COLUMNA
    ]

    # 🧠 Nombre de hoja = ACUERDO MARCO
    nombre_hoja = acuerdo_seleccionado.strip()[:31]  # Excel máximo 31 caracteres

    # =========================================
    # 📂 CARGAR O CREAR HOJA DEL ACUERDO
    # =========================================
    from openpyxl import load_workbook

    if os.path.exists(ruta_excel):
        try:
            wb = load_workbook(ruta_excel)

            if nombre_hoja in wb.sheetnames:

                try:
                    df = pd.read_excel(ruta_excel, sheet_name=nombre_hoja)

                    # 🔴 SI LA HOJA ESTÁ VACÍA
                    if df.empty and len(df.columns) == 0:
                        print("⚠ Hoja vacía detectada — creando columnas")
                        df = pd.DataFrame(columns=columnas)

                    # 🔧 ASEGURAR QUE TODAS LAS COLUMNAS EXISTAN
                    for col in columnas:
                        if col not in df.columns:
                            df[col] = ""

                    df = df[columnas]  # ordenar columnas

                    print(f"📂 Hoja '{nombre_hoja}' cargada correctamente.")

                    # =========================================
                    # 🧠 CREAR SET DE CLAVES EXISTENTES
                    # =========================================
                    claves_existentes = set()

                    if not df.empty:

                        for _, r in df.iterrows():

                            clave = (
                                limpiar_valor(r["REQUERIMIENTO"]),
                                limpiar_valor(r["PROFORMA"]),
                                limpiar_valor(r["RUC"]),
                                limpiar_valor(r["CODIGO"]),
                                limpiar_valor(r["FICHA PRODUCTO"])
                            )

                            claves_existentes.add(clave)

                    print("🧠 Claves cargadas desde Excel:", len(claves_existentes))

                    print("\n📋 CLAVES EXISTENTES EN EXCEL:")
                    for c in claves_existentes:
                        print("EXCEL:", c)

                except Exception as e:

                    print("⚠ Error leyendo hoja — recreando estructura:", e)

                    df = pd.DataFrame(columns=columnas)

            else:

                print(f"📄 Hoja '{nombre_hoja}' no existe — creando nueva")

                df = pd.DataFrame(columns=columnas)

        except Exception as e:
            print("⚠ Error leyendo Excel:", e)
            df = pd.DataFrame(columns=columnas)

    else:
        print("📂 Archivo no existe. Se creará nuevo.")
        df = pd.DataFrame(columns=columnas)




        
        


        
        
    pagina_actual = 1
    MAX_PAGINAS = 25  # 🔥 LÍMITE DE SEGURIDAD ANTI LOOP INFINITO

    while pagina_actual <= MAX_PAGINAS:
        print(f"\n🧭 ===== PROCESANDO PÁGINA {pagina_actual} =====")
        
        


        # ==========================================================
        # CAPTURAR FILAS DE LA PÁGINA ACTUAL
        # ==========================================================
        # 🛡️ ESPERAR QUE TABLA TERMINE DE CARGAR
        if not esperar_loader_robusto(driver, 35):
            recuperar_de_freeze(driver, pagina_actual)

        filas = WebDriverWait(driver, 15).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "#tbData .FilaDatos"))
        )

        print(f"➡ Filas encontradas en página {pagina_actual}: {len(filas)}")

        # 🔑 GUARDAR FIRMA DE ESTA PÁGINA
        firma_pagina_actual = obtener_firma_pagina(driver)
        print(f"🔑 Firma página {pagina_actual}: {firma_pagina_actual}")

        # 🛡️ VALIDAR QUE LA TABLA REALMENTE EXISTE (ANTI DOM MUERTO)
        firma_verificacion = obtener_firma_pagina(driver)

        if not firma_verificacion:
            print("💀 TABLA VACÍA O DOM MUERTO — RECUPERANDO")
            recuperar_de_freeze(driver, pagina_actual)
            continue


                
        


        # ==========================================================
        # RECORRER FILAS (TU LÓGICA ORIGINAL)
        # ==========================================================
        idx = 0

        while True:

            # 🛡️ SIEMPRE recapturar filas reales
            if not esperar_loader_robusto(driver, 35):
                recuperar_de_freeze(driver, pagina_actual)

            filas = WebDriverWait(driver, 15).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, "#tbData .FilaDatos"))
            )

            if idx >= len(filas):
                break

            try:
                fila = filas[idx]
            except:
                print("⚠ Fila inválida tras recarga — reintentando misma posición")
                continue

        
            
            try:

                # 🔁 VERIFICAR QUE NO HAYA REGRESADO A OTRA PÁGINA
                if pagina_actual > 1:
                    firma_actual_en_vista = obtener_firma_pagina(driver)

                    if firma_actual_en_vista != firma_pagina_actual:
                        print(f"🔄 La web regresó a otra página. Volviendo a página {pagina_actual}…")
                        ir_a_pagina(driver, pagina_actual)

                        WebDriverWait(driver, 15).until(
                            lambda d: obtener_firma_pagina(d) == firma_pagina_actual
                        )

                        # 🔥 RECARGAR FILAS DESPUÉS DE VOLVER
                        filas = WebDriverWait(driver, 15).until(
                            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "#tbData .FilaDatos"))
                        )

                        if idx >= len(filas):
                            print("⚠ Índice fuera de rango después de recargar página — saltando fila")
                            continue
            
            

                # 🔁 ASEGURAR QUE ESTAMOS EN LA PÁGINA CORRECTA
                ##if pagina_actual > 1:
                ##  ir_a_pagina(driver, pagina_actual)

                try:
                    # 🛡️ ESPERAR QUE TABLA TERMINE DE CARGAR
                    if not esperar_loader_robusto(driver, 35):
                        recuperar_de_freeze(driver, pagina_actual)

                    filas = WebDriverWait(driver, 15).until(
                        EC.presence_of_all_elements_located((By.CSS_SELECTOR, "#tbData .FilaDatos"))
                    )

                    try:
                        fila = filas[idx]
                    except IndexError:
                        print("⚠ Fila ya no existe tras recarga — continuando")
                        continue
                    
                except:
                    print("❌ No se pudo recapturar fila. Continuando…")
                    continue


                print(f"\n----- FILA {idx+1} (PÁG {pagina_actual}) -----")

                celdas = fila.find_elements(By.TAG_NAME, "td")

                try:
                    requerimiento = celdas[0].text.strip()
                    procedimiento = celdas[1].text.strip()
                    proforma      = celdas[4].text.strip()
                    fecha_emision = celdas[5].text.strip()
                    fecha_limite  = celdas[8].text.strip()
                    entidad       = celdas[10].text.strip()
                    ruc           = celdas[11].text.strip()
                except Exception as e:
                    print("❌ Error leyendo columnas:", e)
                    continue

                if not esperar_loader_robusto(driver, 25):
                    recuperar_de_freeze(driver, pagina_actual)
                    
                estabilizar_tabla_antes_modal(driver, pagina_actual)


                if not click_cotizar_robusto(driver, idx, pagina_actual):
                    print("💀 No se pudo hacer click después de varios intentos — saltando fila")
                    continue


                # ==========================================================
                # 🔁 REINTENTO AUTOMÁTICO SI MODAL FALLA
                # ==========================================================

                # ==========================================================
                # 🔁 REINTENTO INTELIGENTE SI MODAL NO ABRE O SE CUELGA
                # ==========================================================

                modal_ok = False

                for intento in range(4):

                    # 🧠 NUEVA TÉCNICA → ESTABILIZAR ANTES DE INTENTAR
                    estabilizar_tabla_antes_modal(driver, pagina_actual)

                    try:
                        WebDriverWait(driver, 6).until(
                            EC.visibility_of_element_located((By.CSS_SELECTOR, "#divItem .item"))
                        )

                        item = driver.find_element(By.CSS_SELECTOR, "#divItem .item")
                        texto = item.text.strip()

                        if len(texto) < 20:
                            raise Exception("Modal colgado")

                        modal_ok = True
                        break

                    except Exception:
                        print(f"⚠ Modal falló o colgado → intento {intento+1}")

                        cerrar_modal_forzado(driver)

                        # 🔁 Esperar tabla estable ANTES de volver a clickear
                        estabilizar_tabla_antes_modal(driver, pagina_actual)

                        try:
                            filas = driver.find_elements(By.CSS_SELECTOR, "#tbData .FilaDatos")
                            boton = filas[idx].find_element(By.CSS_SELECTOR, "button[title='Cotizar']")
                            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", boton)
                            driver.execute_script("arguments[0].click();", boton)
                            time.sleep(1.2)
                        except:
                            print("❌ No se pudo reintentar click")

                if not modal_ok:
                    print("💀 Modal imposible de abrir — saltando fila")
                    continue


                try:
                    
                    item = driver.find_element(By.CSS_SELECTOR, "#divItem .item")
                    
                    
                    # ==========================
                    # 📄 PDF (LINK REAL)
                    # ==========================
                    pdf_link = ""

                    try:
                        pdf_a = item.find_element(
                            By.CSS_SELECTOR,
                            "a[href*='blob.core.windows.net']"
                        )
                        pdf_link = pdf_a.get_attribute("href").strip()
                    except:
                        pdf_link = ""


                    # 🧪 DEBUG
                    print(f"📄 PDF EXTRAÍDO: {pdf_link}")


                    imagen_link = obtener_url_imagen_desde_pdf(pdf_link)

                    print(f"🖼 IMAGEN GENERADA: {imagen_link}")

                    # ==========================
                    # 💰 SUBTOTAL REAL (TABLA ENTREGAS)
                    # ==========================
                    entrega = driver.find_element(By.CSS_SELECTOR, "#divEntregas .item")

                    celdas_entrega = entrega.find_elements(By.CSS_SELECTOR, "div[role='cell']")

                    # Última columna = Sub Total
                    subtotal = celdas_entrega[-1].text.strip()

                    # 🧪 DEBUG CLARO
                    print(f"💰 SUBTOTAL EXTRAÍDO: {subtotal}")


                    entrega = driver.find_element(By.CSS_SELECTOR, "#divEntregas .item")

                    producto = item.find_elements(By.CLASS_NAME, "flex-row")[0].text.strip()
                    ficha = item.find_elements(By.CLASS_NAME, "flex-row")[2].text.strip()
                    cantidad = item.find_elements(By.CLASS_NAME, "flex-row")[5].text.strip()

                    # 🔍 MARCA DESDE DICCIONARIO
                    marca = detectar_marca(ficha)

                    # 🔑 CÓDIGO REAL DESDE TEXTO DE FICHA
                    codigo = extraer_codigo_desde_ficha(ficha)

                    # ==========================================================
                    # 🔑 CREAR CLAVE ÚNICA DEL REGISTRO
                    # ==========================================================
                    clave_actual = (
                        limpiar_valor(requerimiento),
                        limpiar_valor(proforma),
                        limpiar_valor(ruc),
                        limpiar_valor(codigo),
                        limpiar_valor(ficha)
                    )

                    


        
                    # ==========================================================
                    # 🛑 CONTROL DE DUPLICADOS
                    # ==========================================================
                    if clave_actual in claves_existentes:
                        print("⏭ Registro ya existe en Excel — saltando")
                        
                        cerrar_modal_forzado(driver)  # cerrar modal
                        
                        idx += 1
                        continue
                        

                    # 🧪 DEBUG (opcional pero recomendado)
                    print(f"🔎 MARCA: {marca} | CÓDIGO: {codigo}")


                    direccion = entrega.find_elements(By.CLASS_NAME, "flex-row")[1].text.strip()
                    inicio = entrega.find_elements(By.CLASS_NAME, "flex-row")[2].text.strip()
                    plazo = entrega.find_elements(By.CLASS_NAME, "flex-row")[3].text.strip()
                    fin = entrega.find_elements(By.CLASS_NAME, "flex-row")[4].text.strip()
                except Exception as e:
                    print("⚠️ Error leyendo el modal:", e)
                    try:
                        cerrar_modal_forzado(driver)
                    except:
                        pass
                    continue

                clave_actual = (
                    limpiar_valor(requerimiento),
                    limpiar_valor(proforma),
                    limpiar_valor(ruc),
                    limpiar_valor(codigo),
                    limpiar_valor(ficha)
                   
                )


                # ==========================================================
                # 🛑 CONTROL DE DUPLICADOS
                # ==========================================================
                if clave_actual in claves_existentes:
                    print("🔁 FILA REPETIDA:", clave_actual)

                    cerrar_modal_forzado(driver)

                    idx += 1
                    continue



            
                nueva_fila = {
                    "N°": int(df["N°"].max()) + 1 if not df["N°"].isna().all() else 1,
                    "FECHA_GUARDADO": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "PAGINA": pagina_actual,
                    "FILA_PAGINA": idx + 1,

                    "REQUERIMIENTO": requerimiento,
                    "PROCEDIMIENTO": procedimiento,
                    "PROFORMA": proforma,
                    "FECHA DE EMISION": fecha_emision,
                    "FECHA LIMITE DE COTIZACION": fecha_limite,
                    "ENTIDAD": entidad,
                    "RUC": ruc,
                    "PRODUCTO": producto,
                    "FICHA PRODUCTO": ficha,
                    "MARCA": marca,
                    "CODIGO": codigo,
                    "CANTIDAD": cantidad,
                    "DIRECCION": direccion,
                    "INICIO DE ENTREGA": inicio,
                    "PLAZO MAXI": plazo,
                    "FIN DE ENTREGA": fin,
                    "SUBTOTAL": subtotal,
                    "PDF": pdf_link,
                    "IMAGEN": imagen_link
                }
                df = pd.concat([df, pd.DataFrame([nueva_fila])], ignore_index=True)
                claves_existentes.add(clave_actual)

                # 💾 GUARDAR EN EXCEL
                with pd.ExcelWriter(
                    ruta_excel,
                    engine="openpyxl",
                    mode="a" if os.path.exists(ruta_excel) else "w",
                    if_sheet_exists="replace"
                ) as writer:

                    df.to_excel(writer, sheet_name=nombre_hoja, index=False)

                print("✅ GUARDADO OK")

    

                try:
                    cerrar_modal_forzado(driver)
                    time.sleep(0.4)
                except:
                    pass
            
            
            except Exception as e:
                print(f"💥 Error crítico en fila {idx+1} página {pagina_actual}: {e}")
                cerrar_modal_forzado(driver)
                idx += 1   # 🔥 AVANZA DE FILA SÍ O SÍ
                continue
            
            idx += 1
            
        # ======================================================
        # 🧠 FORZAR POSICIÓN REAL ANTES DE PAGINAR
        # ======================================================

        print("🧠 Verificando posición real antes de paginar...")

        if pagina_actual > 1:
            ir_a_pagina(driver, pagina_actual)

        time.sleep(0.8)

        # ======================================================
        # 🧠 PAGINACIÓN CON DETECCIÓN DE FREEZE
        # ======================================================

        exito = ir_a_siguiente_pagina_robusto(driver, pagina_actual)



        if not exito:
            print("🏁 No se pudo continuar — fin real")
            break

        pagina_actual += 1
        print(f"➡ Página {pagina_actual} confirmada")
        
        # ======================================================
        # 🛡️ ESPERAR CARGA REAL DE LA TABLA (CRÍTICO)
        # ======================================================

        print("⏳ Esperando carga completa de la tabla...")

        if not esperar_loader_robusto(driver, 40):
            print("💀 Freeze detectado tras paginar — recuperando")
            recuperar_de_freeze(driver, pagina_actual)

        time.sleep(1)

    print("\n✅ Se realizó la extracción de datos — TODAS LAS PÁGINAS")
    
    
    
    
    
    
def extraer_restringir(driver, acuerdo_seleccionado):

    restricciones_hechas = 0
    driver.set_page_load_timeout(60)   # ← AQUÍ EXACTAMENTE
    import os
    import time
    import pandas as pd
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC


    # ==========================================================
    # 🧠 SISTEMA ROBUSTO DE CARGA
    # ==========================================================

    def esperar_loader(driver, timeout=15):

        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC

        try:
            WebDriverWait(driver, timeout).until_not(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, ".loading, .spinner, .k-loading-mask")
                )
            )
            return True
        except:
            return False


    def esperar_loader_robusto(driver, timeout_total=40):

        import time
        from selenium.webdriver.common.by import By

        inicio = time.time()

        while True:

            filas = driver.find_elements(By.CSS_SELECTOR, "#tbData .FilaDatos")

            loaders = driver.find_elements(
                By.CSS_SELECTOR,
                ".k-loading-mask:not([style*='display: none'])"
            )

            if len(filas) == 0:
                time.sleep(1)
                continue

            if not loaders and len(filas) > 0:
                return True

            if loaders and len(filas) > 0:

                try:
                    driver.execute_script("""
                        var l = document.querySelector('.k-loading-mask');
                        if(l){ l.style.display='none'; }
                    """)
                except:
                    pass

                return True

            if time.time() - inicio > timeout_total:
                return False

            time.sleep(0.5)


    def recuperar_de_freeze(driver, pagina_actual):

        import time
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC

        print("♻ RECUPERANDO SISTEMA")

        try:
            driver.execute_script("window.stop();")
        except:
            pass

        try:
            driver.refresh()
            time.sleep(2)
        except:
            pass

        WebDriverWait(driver, 40).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "#tbData"))
        )

        esperar_loader_robusto(driver, 40)

        ir_a_pagina(driver, pagina_actual)

        print("✅ SISTEMA RECUPERADO")


    def estabilizar_tabla_antes_modal(driver, pagina_actual):

        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        import time

        if not esperar_loader_robusto(driver, 30):
            recuperar_de_freeze(driver, pagina_actual)

        WebDriverWait(driver, 15).until(
            EC.presence_of_all_elements_located(
                (By.CSS_SELECTOR, "#tbData .FilaDatos")
            )
        )

        time.sleep(0.6)


    def cerrar_modal_forzado(driver):

        from selenium.webdriver.common.by import By
        import time

        try:

            boton = driver.find_element(By.ID, "btnCerraPopupCotizacion")

            driver.execute_script(
                "arguments[0].click();",
                boton
            )

            time.sleep(1)

            esperar_loader_robusto(driver, 30)

            return True

        except:
            return False


    def click_cotizar_robusto(driver, idx, pagina_actual, max_intentos=5):

        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        import time

        for intento in range(max_intentos):

            try:

                estabilizar_tabla_antes_modal(driver, pagina_actual)

                filas = WebDriverWait(driver, 15).until(
                    EC.presence_of_all_elements_located(
                        (By.CSS_SELECTOR, "#tbData .FilaDatos")
                    )
                )

                fila = filas[idx]

                boton = fila.find_element(
                    By.CSS_SELECTOR,
                    "button[title='Cotizar']"
                )

                driver.execute_script(
                    "arguments[0].scrollIntoView({block:'center'});",
                    boton
                )

                time.sleep(0.3)

                driver.execute_script(
                    "arguments[0].click();",
                    boton
                )

                return True

            except Exception:

                esperar_loader(driver)

                ir_a_pagina(driver, pagina_actual)

                time.sleep(1)

        return False


    
    
    # ==========================================================
    # 🧠 DICCIONARIO DE MARCAS (LIMPIEZA + CEREALES + BEBIDAS)
    # ==========================================================
    MARCAS_DICC = [
        # Limpieza / Industrial / Hogar
        "W&M SIRYCATA","ITZEL","D-SANIZ","PÓMAC","MILLENNIUM","KYSER",
        "SERCONSLIMP","COVENANT","C3 P","SUPREMIO","ULTRA+","RINRI",
        "RAYMI","APOLO","BELONA","ISIS","ECOKASA","PARILU","SINCHY",
        "CIRCE","DELSA","REGGIO","D SALOME","JARVIS","CICLÓN",
        "ECOLIMPIA","SILGAL","DIONNE","SUAVISSA","PILLKO","AYCER",
        "MAXCER","ANAX","PLASTIC STEEL FORJA","STEEL FORJA","DICALI",
        "AITY","+ASEO","LAMOSA CLEAN","H TOOLS","DQ DELQUIMS",
        "NOELIA","CELESTINA","JAWASS","VIANFORTPRO","BIODECOR",
        "BLANUX","MUCHICK","YOREL","PROSERLIM","ZENCLEAN",
        "GOOD CLEANER","TACLLA","EBRIEL","JAZAY","VICRO","LIFAL",
        "RAGNAR","SALPE","REY PLAST","SMARTPLAST","CERCOR","CETOOL",
        "ATOJ","BUGUI","ESTRELLA DEL NORTE","DARYAL","DULQUI",
        "GP GRIUPOLY","OSCCONTA","INDUBRILL","KUELAP","RHOMANSA",
        "INSOMED","NEW KRAL","ESE","MAXIMASS","WYPALL","JOGRANSA",
        "DIMAGSA","PROLIM","BOMELSA","ECOSOFT","MAPIALE","TEXTILES ECOKASA",
        "THN","DARYZA","SOLMATIC","DOÑA MARIA","AGLAB","ECONORACKS",
        "FABARLI","DALHI","MAGINSA","MAFA","SCOTT","LIMCOFER",
        "WHUAYRA","INVEMATT","YANNICK","INVERCOM PERU","SFOLL",
        "DAYR","RS PASSIONES TEX","NEW AYMAX","PULIZIA","PERCUS",
        "T&R CLEANER","DERMA PRO","FACIL","HOUSE LIVING","MVILLEGAS",
        "BASA","AIME","JHANSE","ELITE","ARUBBA","WIPE MASTER",
        "ARIESS","MEGACOM","YERICO","NORT COLTON","JEKMAYCAN",
        "TOOLBOX","COMPAKTO1","SCARLETT","INSTITUCIONAL SUPER",
        "ELITE PROFESSIONAL","MELISSA","GLAX","+PRO","HI LIMP",
        "SUMAC","TOILÉ","THAILER","3Q PLASTICS","FONLEA",
        "CERMAX CLEAN","3Q COTTONS","CAPACMAYO","FBK PERÚ",
        "VIRUTEX","LACTISOFT","ALESSI","MAXTIC","PRL PARILU EXPRESS",
        "MOTITA","LEONSOL","ALICAF","LYNO","ELLIA","PARACAS",
        "CISNE","MASSIA","LA FOQUITA","PROLIMSO","CHACON","F FERNELLY",
        "SIAL","GRAFI PAPEL","J&R STEEL CP","SANIT","FONCHY",
        "FAMILY DOCTOR","BIGNER","MAYA","TUINIES","SUPER",
        "WEST MICROSAFE","AYAX","KLEENEX","HANDEEL","BRILLOL'S",
        "DERQUSA","Q MASTER","SUAVE","VIBALCA","PETALO","BUBBLE",
        
        #ACCESORIOS DOMÉSTICOS
        "LLANKAQ",
        

        # Cereales / Abarrotes
        "KASQUI","KOMILON","COSTEÑO","MENESTRERO","DEL CIELO",
        "PAISANA","CERRO GRANDE","JAPRIM","MOLINO ROJO","COSTEÑITO",
        "OLLITA","FORTILIFE","PALMA REAL","DEL NORTE","PURA CAÑA",
        "HOJA REDONDA","DEL HOGAR","DOÑA RUFI",

        # Bebidas
        "VITALIA","CIELO"
    ]
    
    
    
    # ==========================================================
    # 🚫 MARCAS RESTRINGIDAS (NUEVO)
    # ==========================================================
    MARCAS_RESTRINGIDAS = {
        "ECOLIMPIA","DICALI","CICLÓN","ANAX",
        "INDUBRILL","FACIL","PARILU","PARILÚ","INSOMED",
        "BLANUX","ATOJ","DULQUI","RAYMI",
        "AIME","DELQUIMS","KUELAP","STEEL FORJA",
        "Q MASTER","ECO CLEAN","PÓMAC","ESTRELLA DEL NORTE",
        "REGGIO","DERMA PRO","TOOLS","ISIS",
        "MUCHICK","APOLO","SINCHY","CISNE",
        "DIONNE","VICRO","TACLLA","ZENCLEAN",
        "VIANFORTPRO","BELONA","BOMELSA","PROTEGE",
        "AYCER","GOOD CLEANER","SILGAL",
        "MELISSA","BRICEL","MAXCER","GRIUPOLY",
        "W&M SIRYCATA","CERCOR","CETOOL","JARVIS",
        "REYSER","CIRCE","LIFAL","EL PÁRAMO PERÚ",
        "THN","ITZEL","WADA"
    }

    
    def detectar_marca(ficha):
        texto = ficha.upper()
        import re

        # 1️⃣ Buscar "MARCA: XXXX" pero validar contra diccionario
        match = re.search(r"MARCA\s*:\s*([A-Z0-9 +&'.-]+)", texto)
        if match:
            posible = match.group(1).strip()

            # devolver SOLO la marca válida del diccionario
            for marca in sorted(MARCAS_DICC, key=len, reverse=True):
                if posible.startswith(marca):
                    return marca

        # 2️⃣ Fallback: buscar marca directa en el texto
        for marca in sorted(MARCAS_DICC, key=len, reverse=True):
            if f" {marca} " in f" {texto} ":
                return marca

        return ""


    

    def extraer_codigo_desde_ficha(ficha):
        import re

        if not ficha:
            return ""

        # ===============================
        # 1️⃣ NORMALIZACIÓN FUERTE
        # ===============================
        texto = ficha.upper()
        texto = texto.replace("–", "-").replace("—", "-")
        texto = re.sub(r'\s+', ' ', texto).strip()

        # ===============================
        # 2️⃣ LIMPIEZA DE BASURA SEGURA
        # ===============================
        # Quitar EAN / CUBSO largos
        texto = re.sub(r'\b\d{13,15}\b', ' ', texto)

        # Quitar rangos y porcentajes (7.50 - 8.50, 20 - 33.00)
        texto = re.sub(r'\b\d+(\.\d+)?\s*-\s*\d+(\.\d+)?\b', ' ', texto)

        # Quitar tiempos
        texto = re.sub(r'\b\d+(\.\d+)?\s*(MESES|AÑOS|AÑO)\b', ' ', texto)

        # ===============================
        # 3️⃣ TRABAJAR SOLO CON EL FINAL
        # ===============================
        corte = int(len(texto) * 0.65)
        texto_final = texto[corte:]

        # ===============================
        # 4️⃣ PATRONES ROBUSTOS (ORDEN IMPORTA)
        # ===============================
        patrones = [

            # 🔥 Con símbolos + (RECOGEDOR+FILO+369)
            r'\b[A-Z0-9]+(?:\+[A-Z0-9]+){1,}\b',

            # 🔥 Múltiples guiones (TACHO-C-1201SN)
            r'\b[A-Z0-9]+(?:-[A-Z0-9.%]+){2,}\b',

            # 🔥 Guión simple completo (RE23-A, SHP7.5-20LT, 403-4LT)
            r'\b[A-Z0-9]{2,}\s*-\s*[A-Z0-9.]+\b',

            # 🔥 Marca + número (EBRIEL 732, DARYZA 31422, 1119024 VIRUTEX)
            r'\b[A-Z]{3,}\s+\d{3,8}\b|\b\d{3,8}\s+[A-Z]{3,}\b',

            # 🔥 Alfanumérico compacto (ERPLAZ, HL20L7.5%)
            r'\b[A-Z]{2,}\d+[A-Z0-9.%]*\b',

            # 🔥 Numérico puro válido (13120043, 300107, 31423)
            r'\b\d{5,8}\b',
        ]

        # ===============================
        # 5️⃣ BUSCAR DESDE EL FINAL
        # ===============================
        for patron in patrones:
            matches = list(re.finditer(patron, texto_final))
            if matches:
                codigo = matches[-1].group(0).strip()
                return codigo

        return ""

    # ==========================================================
    # 🚀 FUNCIÓN ULTRA ROBUSTA PARA IR A UNA PÁGINA
    # ==========================================================
    def ir_a_pagina(driver, numero_objetivo):

        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        import time

        print(f"🎯 Navegando a página {numero_objetivo}")

        try:

            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.ID, "divPaginacion"))
            )

            # --------------------------------------------------
            # 1️⃣ DETECTAR PÁGINA ACTUAL REAL
            # --------------------------------------------------
            try:
                pagina_actual_real = int(obtener_pagina_activa(driver))
            except:
                pagina_actual_real = 1

            if pagina_actual_real == numero_objetivo:
                print("🟢 Ya estamos en la página correcta")
                return True

            # --------------------------------------------------
            # 2️⃣ INTENTAR CLICK DIRECTO
            # --------------------------------------------------
            boton_directo = driver.find_elements(
                By.XPATH,
                f"//div[@id='divPaginacion']//input[@value='{numero_objetivo}']"
            )

            if boton_directo:

                driver.execute_script(
                    "arguments[0].scrollIntoView({block:'center'});",
                    boton_directo[0]
                )

                driver.execute_script(
                    "arguments[0].click();",
                    boton_directo[0]
                )

                WebDriverWait(driver, 25).until(
                    lambda d: obtener_pagina_activa(d) == str(numero_objetivo)
                )

                esperar_loader_robusto(driver)

                print(f"✅ Página {numero_objetivo} alcanzada directo")
                return True

            # --------------------------------------------------
            # 3️⃣ BUSCAR ANCLA VISIBLE
            # --------------------------------------------------
            botones = driver.find_elements(
                By.XPATH,
                "//div[@id='divPaginacion']//input[@type='button']"
            )

            numeros = []

            for b in botones:
                try:
                    val = int(b.get_attribute("value"))
                    numeros.append(val)
                except:
                    pass

            if not numeros:
                print("💀 No se encontraron botones de página")
                return False

            ancla = max([n for n in numeros if n < numero_objetivo], default=None)

            if not ancla:
                print("💀 No hay ancla válida")
                return False

            print(f"⚓ Usando ancla {ancla}")

            boton_ancla = driver.find_element(
                By.XPATH,
                f"//div[@id='divPaginacion']//input[@value='{ancla}']"
            )

            driver.execute_script(
                "arguments[0].scrollIntoView({block:'center'});",
                boton_ancla
            )

            driver.execute_script(
                "arguments[0].click();",
                boton_ancla
            )

            WebDriverWait(driver, 25).until(
                lambda d: obtener_pagina_activa(d) == str(ancla)
            )

            esperar_loader_robusto(driver)

            time.sleep(1)

            # --------------------------------------------------
            # 4️⃣ AJUSTE FINO CON "SIGUIENTE"
            # --------------------------------------------------
            pagina_actual_real = ancla

            while pagina_actual_real < numero_objetivo:

                boton_siguiente = WebDriverWait(driver, 15).until(
                    EC.element_to_be_clickable((
                        By.XPATH,
                        "//div[@id='divPaginacion']//input[contains(@value,'Siguiente')]"
                    ))
                )

                driver.execute_script(
                    "arguments[0].scrollIntoView({block:'center'});",
                    boton_siguiente
                )

                driver.execute_script(
                    "arguments[0].click();",
                    boton_siguiente
                )

                WebDriverWait(driver, 25).until(
                    lambda d: obtener_pagina_activa(d) == str(pagina_actual_real + 1)
                )

                esperar_loader_robusto(driver)

                pagina_actual_real += 1

                print(f"➡ Avanzó a página {pagina_actual_real}")

                time.sleep(0.6)

            print(f"✅ Página {numero_objetivo} alcanzada correctamente")

            return True

        except Exception as e:

            print(f"💀 Error navegando a página {numero_objetivo}: {e}")

            return False



            
            
            
    def ir_a_siguiente_pagina_robusto(driver, pagina_actual):
        print("\n🔎 Intentando ir a la siguiente página...")
        return ir_a_pagina(driver, pagina_actual + 1)





    # ==========================================================
    # 🔑 FUNCIÓN FIRMA DE PÁGINA (CONTROL DE PAGINACIÓN)
    # ==========================================================
    def obtener_firma_pagina(driver):
        try:
            filas = WebDriverWait(driver, 20).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, "#tbData .FilaDatos"))
            )

            if not filas:
                return "SIN_FILAS"
            return filas[0].find_element(By.CSS_SELECTOR, "td.primerTd").text.strip()
        except:
            return "ERROR_FIRMA"
        
        
    def obtener_pagina_activa(driver):
        try:
            return driver.find_element(
                By.CSS_SELECTOR,
                "#divPaginacion input.NavegaActivo"
            ).get_attribute("value")
        except:
            return "0"




    print("\n📊 INICIANDO EXTRACCIÓN…")

    try:
        driver.execute_script("document.body.style.zoom='75%'")
    except:
        pass

    ruta_excel = ruta_recurso("assets/RESTRI.xlsx")
    
    
    # 🧠 Nombre de hoja = ACUERDO MARCO
    nombre_hoja = acuerdo_seleccionado.strip()[:31]


    columnas = [
        "N°", "REQUERIMIENTO", "PROCEDIMIENTO", "PROFORMA",
        "FECHA DE EMISION", "FECHA LIMITE DE COTIZACION",
        "ENTIDAD", "RUC", "PRODUCTO", "FICHA PRODUCTO",
        "MARCA", "CODIGO", "CANTIDAD", "DIRECCION",
        "INICIO DE ENTREGA", "PLAZO MAXI", "FIN DE ENTREGA",
        "SUBTOTAL", "PDF"
    ]

    # ==========================================================
    # CARGAR EXCEL
    # ==========================================================
    # ==========================================================
    # 📂 CARGAR O CREAR HOJA DEL ACUERDO (RESTRI)
    # ==========================================================
    if os.path.exists(ruta_excel):
        try:
            df = pd.read_excel(ruta_excel, sheet_name=nombre_hoja)
            df = df.fillna("").astype(str).applymap(lambda x: x.strip())
            print(f"📂 Hoja '{nombre_hoja}' cargada en RESTRI.xlsx")
        except:
            print(f"📄 Hoja '{nombre_hoja}' no existe en RESTRI.xlsx. Se creará.")
            df = pd.DataFrame(columns=columnas)
    else:
        print("📂 RESTRI.xlsx no existe. Se creará nuevo.")
        df = pd.DataFrame(columns=columnas)
        
        


    # ==========================================================
    # 🧠 CONTROL DE CLAVES EXISTENTES
    # ==========================================================
    claves_existentes = set()
            

    pagina_actual = 1

    while True:
        print(f"\n🧭 ===== PROCESANDO PÁGINA {pagina_actual} =====")
        
        

        if not esperar_loader_robusto(driver, 40):
            recuperar_de_freeze(driver, pagina_actual)
        # ==========================================================
        # CAPTURAR FILAS DE LA PÁGINA ACTUAL
        # ==========================================================
        filas = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "#tbData .FilaDatos"))
        )
        print(f"➡ Filas encontradas en página {pagina_actual}: {len(filas)}")

        # 🔑 GUARDAR FIRMA DE ESTA PÁGINA
        firma_pagina_actual = obtener_firma_pagina(driver)
        print(f"🔑 Firma página {pagina_actual}: {firma_pagina_actual}")
        
      


                
        


        # ==========================================================
        # RECORRER FILAS (TU LÓGICA ORIGINAL)
        # ==========================================================
        # ==========================================================
        # RECORRER FILAS (VERSIÓN CONTROLADA)
        # ==========================================================
        total_filas = len(filas)   # 🔥 CONGELAR TOTAL REAL

        idx = 0
        while idx < total_filas:


            
            # 🔁 VERIFICAR QUE NO HAYA REGRESADO A PÁGINA 1
            # 🚀 VERIFICACIÓN SOLO DESDE PÁGINA 2 EN ADELANTE


            
            

            # 🔁 ASEGURAR QUE ESTAMOS EN LA PÁGINA CORRECTA
            ##if pagina_actual > 1:
              ##  ir_a_pagina(driver, pagina_actual)

            try:

                if not esperar_loader_robusto(driver, 40):
                    recuperar_de_freeze(driver, pagina_actual)
                filas = WebDriverWait(driver, 10).until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, "#tbData .FilaDatos"))
                )
                fila = filas[idx]
            except:
                print("❌ No se pudo recapturar fila. Continuando…")
                idx += 1
                continue


            print(f"\n----- FILA {idx+1} (PÁG {pagina_actual}) -----")

            celdas = fila.find_elements(By.TAG_NAME, "td")
            
            
            
            # ==========================================================
            # 🔴 VALIDAR INDICADOR ROJO ANTES DE COTIZAR
            # ==========================================================
            try:
                indicador_texto = celdas[13].text.strip().upper()
                if "ROJO" not in indicador_texto:
                    
                    print(f"⏭ FILA {idx+1}: Indicador NO rojo ({indicador_texto})")

                    # =====================================================
                    # 🔥 SI ES LA ÚLTIMA FILA → HACER CLICK FANTASMA
                    # =====================================================
                    if idx == total_filas - 1:

                        print("🟢 Última fila verde → ejecutando click fantasma robusto...")

                        try:

                            if not click_cotizar_robusto(driver, idx, pagina_actual):
                                print("⚠ No se pudo ejecutar click fantasma")
                                idx += 1
                                continue

                            # esperar modal
                            WebDriverWait(driver, 10).until(
                                EC.visibility_of_element_located((By.ID, "btnCerraPopupCotizacion"))
                            )

                            time.sleep(1)  # solo visual

                            # cerrar modal
                            cerrar_modal_forzado(driver)

                            # esperar que vuelva la tabla
                            WebDriverWait(driver, 20).until(
                                EC.presence_of_all_elements_located((By.CSS_SELECTOR, "#tbData .FilaDatos"))
                            )
                            time.sleep(1.5)

                            print("✅ Click fantasma ejecutado correctamente")

                        except Exception as e:
                            print("⚠ Error en click fantasma:", e)

                    idx += 1
                    continue


                else:
                    print(f"🔴 FILA {idx+1}: Indicador ROJO detectado → Entrando a cotizar")
                    
                   


            except Exception as e:
                print(f"⚠ No se pudo validar indicador en fila {idx+1}: {e}")
                idx += 1   # 👈 AGREGAR
                continue


            try:
                requerimiento = celdas[0].text.strip()
                procedimiento = celdas[1].text.strip()
                proforma      = celdas[4].text.strip()
                fecha_emision = celdas[5].text.strip()
                fecha_limite  = celdas[8].text.strip()
                entidad       = celdas[10].text.strip()
                ruc           = celdas[11].text.strip()
            except Exception as e:
                print("❌ Error leyendo columnas:", e)
                idx += 1   # 👈 AGREGAR
                continue

            try:
                if not click_cotizar_robusto(driver, idx, pagina_actual):
                    print("❌ No se pudo abrir el modal")
                    idx += 1
                    continue

            except:
                print("❌ Error al hacer click en Cotizar")
                idx += 1   # 👈 AGREGAR
                continue

            try:
                WebDriverWait(driver, 15).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, "#divItem .item"))
                )
            except:
                print("❌ Modal no apareció")
                idx += 1   # 👈 AGREGAR
                continue

            try:
                
                item = driver.find_element(By.CSS_SELECTOR, "#divItem .item")
                
                
                # ==========================
                # 📄 PDF (LINK REAL)
                # ==========================
                pdf_link = ""

                try:
                    pdf_a = item.find_element(
                        By.CSS_SELECTOR,
                        "a[href*='blob.core.windows.net']"
                    )
                    pdf_link = pdf_a.get_attribute("href").strip()
                except:
                    pdf_link = ""


                # 🧪 DEBUG
                print(f"📄 PDF EXTRAÍDO: {pdf_link}")

                # ==========================
                # 💰 SUBTOTAL REAL (TABLA ENTREGAS)
                # ==========================
                entrega = driver.find_element(By.CSS_SELECTOR, "#divEntregas .item")

                celdas_entrega = entrega.find_elements(By.CSS_SELECTOR, "div[role='cell']")

                # Última columna = Sub Total
                subtotal = celdas_entrega[-1].text.strip()

                # 🧪 DEBUG CLARO
                print(f"💰 SUBTOTAL EXTRAÍDO: {subtotal}")


                entrega = driver.find_element(By.CSS_SELECTOR, "#divEntregas .item")

                producto = item.find_elements(By.CLASS_NAME, "flex-row")[0].text.strip()
                ficha = item.find_elements(By.CLASS_NAME, "flex-row")[2].text.strip()
                cantidad = item.find_elements(By.CLASS_NAME, "flex-row")[5].text.strip()

                # 🔍 MARCA DESDE DICCIONARIO
                marca = detectar_marca(ficha)
                codigo = extraer_codigo_desde_ficha(ficha)

                print("\n==============================")
                print(f"📄 FICHA LEÍDA:")
                print(f"{ficha}")
                print("------------------------------")

                if marca:
                    print(f"🔍 Marca detectada: {marca}")
                else:
                    print("⚠️ No se detectó marca")

                print("🔎 Comparando contra marcas restringidas...")

                if marca in MARCAS_RESTRINGIDAS:
                    print(f"🚫 MARCA RESTRINGIDA ENCONTRADA: {marca}")
                    marca_restringida = True
                else:
                    print(f"✅ Marca NO restringida: {marca}")
                    marca_restringida = False

                print("==============================\n")
                
                
                # ==========================================================
                # 🚫 ACCIÓN SI MARCA ES RESTRINGIDA
                # ==========================================================
                if marca_restringida:
                    
                    print("🚫 Iniciando proceso de RESTRICCIÓN AUTOMÁTICA...")

                    try:
                        # ======================================================
                        # 1️⃣ CLICK EN BOTÓN RESTRINGIR
                        # ======================================================
                        boton_restringir = WebDriverWait(driver, 15).until(
                            EC.element_to_be_clickable((By.ID, "btnRestringirCotizacion"))
                        )

                        driver.execute_script("arguments[0].click();", boton_restringir)
                        time.sleep(2)  # ⬅ pausa para ver el click
                        print("✅ Click en 'Restringir' correcto")

                        # ======================================================
                        # 2️⃣ CLICK EN CONFIRMACIÓN "SI" (VERSIÓN DEFINITIVA REAL)
                        # ======================================================

                        print("⏳ Esperando botón 'Si' activo...")

                        # Esperar que el botón exista en el DOM
                        WebDriverWait(driver, 15).until(
                            EC.presence_of_element_located((
                                By.CSS_SELECTOR,
                                "#confirmationDialog[style*='display: block'] a.btn-success[data-apply='confirmation']"
                            ))
                        )

                        # Click directo por JavaScript puro
                        driver.execute_script("""
                        var btn = document.querySelector(
                            "#confirmationDialog[style*='display: block'] a.btn-success[data-apply='confirmation']"
                        );
                        if(btn){
                            btn.click();
                        }
                        """)
                        time.sleep(2)  # ⬅ pausa para ver el click

                        print("✅ Click en 'Si' ejecutado correctamente")

                        # Esperar que el popover desaparezca
                        WebDriverWait(driver, 15).until(
                            EC.invisibility_of_element_located((By.ID, "confirmationDialog"))
                        )

                        print("✅ Popover cerrado correctamente")
                        

                        # ======================================================
                        # 4️⃣ ESPERAR MODAL DE RESTRICCIÓN
                        # ======================================================

                        print("⏳ Esperando modal de restricción...")

                        WebDriverWait(driver, 15).until(
                            EC.visibility_of_element_located((By.ID, "txtRestringirMotivo"))
                        )
                        time.sleep(0.5)  # ⬅ pequeña pausa para visualización

                        print("✅ Modal de restricción visible")

                        # ======================================================
                        # 5️⃣ SELECCIONAR CAUSAL
                        # ======================================================

                        from selenium.webdriver.support.ui import Select

                        select_causal = WebDriverWait(driver, 15).until(
                            EC.element_to_be_clickable((By.ID, "txtRestringirMotivo"))
                        )

                        select = Select(select_causal)
                        select.select_by_value("POR INDICADOR SEMAFORO")
                        
                        time.sleep(2)

                        print("✅ Causal seleccionada: INDICADOR SEMÁFORO")

                        time.sleep(0.5)

                        # ======================================================
                        # 6️⃣ CLICK EN GUARDAR
                        # ======================================================

                        boton_guardar = WebDriverWait(driver, 15).until(
                            EC.element_to_be_clickable((By.ID, "btnGuardarRestringir"))
                        )

                        driver.execute_script("arguments[0].click();", boton_guardar)
                        time.sleep(2)

                        print("💾 Restricción guardada correctamente")
                        print("🚫 PROCESO COMPLETO FINALIZADO\n")

                        restricciones_hechas += 1

                        time.sleep(1)
                        
                        
 
                        # ======================================================
                        # 7️⃣ CERRAR MODAL DE RESTRICCIÓN
                        # ======================================================
                        print("⏳ Esperando botón 'Cerrar' del modal...")

                        boton_cerrar_restriccion = WebDriverWait(driver, 15).until(
                            EC.element_to_be_clickable((
                                By.CSS_SELECTOR,
                                "button.btn.btn-primary.btn_cerrar"
                            ))
                        )
                        driver.execute_script("arguments[0].click();", boton_cerrar_restriccion)

                        print("✅ Modal de restricción cerrado")

                        # Esperar que el select desaparezca (confirma que cerró)
                        WebDriverWait(driver, 15).until(
                            EC.invisibility_of_element_located((By.ID, "txtRestringirMotivo"))
                        )

                        print("✅ Confirmado: modal completamente cerrado")
                        
                        
                        

                        # 🔹 VOLVER A ZOOM 80%
                        time.sleep(2.2)  # pequeño delay antes de aplicar zoom
                        driver.execute_script("document.body.style.zoom='75%'")
                        print("🔎 Zoom restablecido al 75%")
                        time.sleep(5)  # pequeño delay antes de aplicar zoom

                        # 🔁 RECARGAR TABLA PARA CONTINUAR CON OTRAS MARCAS
                        print("🔄 Recargando tabla después de restricción...")
                        driver.execute_script("location.reload()")
                        time.sleep(5)

                        # esperar que la tabla vuelva a aparecer
                        WebDriverWait(driver, 20).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, "#tbData"))
                        )

                        print("✅ Tabla recargada correctamente")
                                                
                        
 
                    except Exception as e:
                        print(f"❌ ERROR EN PROCESO DE RESTRICCIÓN: {e}")


                direccion = entrega.find_elements(By.CLASS_NAME, "flex-row")[1].text.strip()
                inicio = entrega.find_elements(By.CLASS_NAME, "flex-row")[2].text.strip()
                plazo = entrega.find_elements(By.CLASS_NAME, "flex-row")[3].text.strip()
                fin = entrega.find_elements(By.CLASS_NAME, "flex-row")[4].text.strip()
            except Exception as e:
                print("⚠️ Error leyendo el modal:", e)
                try:
                    cerrar_modal_forzado(driver)
                    
                    WebDriverWait(driver, 20).until(
                        EC.invisibility_of_element_located((By.ID, "btnCerraPopupCotizacion"))
                    )

                    time.sleep(1.5)
                except:
                    pass
                idx += 1   # 👈 AGREGAR
                continue

            clave = (requerimiento, proforma, ruc, codigo, ficha)
            
            if "REQUERIMIENTO" not in df.columns:
                print("🚨 DataFrame corrupto — recreando columnas")
                df = pd.DataFrame(columns=columnas)

            existe = df[
                (df["REQUERIMIENTO"] == clave[0]) &
                (df["PROFORMA"] == clave[1]) &
                (df["RUC"] == clave[2]) &
                (df["CODIGO"] == clave[3]) &
                (df["FICHA PRODUCTO"] == clave[4])
            ]

            # ✅ SOLO GUARDAR FILAS CON MARCA RESTRINGIDA
            if marca_restringida and existe.empty:
                nueva_fila = {
                    "N°": len(df) + 1,
                    "REQUERIMIENTO": requerimiento,
                    "PROCEDIMIENTO": procedimiento,
                    "PROFORMA": proforma,
                    "FECHA DE EMISION": fecha_emision,
                    "FECHA LIMITE DE COTIZACION": fecha_limite,
                    "ENTIDAD": entidad,
                    "RUC": ruc,
                    "PRODUCTO": producto,
                    "FICHA PRODUCTO": ficha,
                    "MARCA": marca,
                    "CODIGO": codigo,
                    "CANTIDAD": cantidad,
                    "DIRECCION": direccion,
                    "INICIO DE ENTREGA": inicio,
                    "PLAZO MAXI": plazo,
                    "FIN DE ENTREGA": fin,
                    "SUBTOTAL": subtotal,
                    "PDF": pdf_link
                }
                df = pd.concat([df, pd.DataFrame([nueva_fila])], ignore_index=True)

                claves_existentes.add(clave)

            try:
                cerrar_modal_forzado(driver)

                WebDriverWait(driver, 20).until(
                    EC.invisibility_of_element_located((By.ID, "btnCerraPopupCotizacion"))
                )

                # 🔥 ESPERAR QUE EL GRID CARGUE
                WebDriverWait(driver, 25).until(
                    lambda d: obtener_firma_pagina(d) != "ERROR_FIRMA"
                )

                time.sleep(1.5)

                # ======================================================
                # 🧠 DETECTAR SI LA WEB REGRESÓ A PÁGINA 1
                # ======================================================
                pagina_real = obtener_pagina_activa(driver)

                if str(pagina_real) != str(pagina_actual):
                    print(f"⚠ La web regresó a página {pagina_real} — Reposicionando a {pagina_actual}")

                    ir_a_pagina(driver, pagina_actual)

                    WebDriverWait(driver, 20).until(
                        lambda d: obtener_pagina_activa(d) == str(pagina_actual)
                    )

                    time.sleep(1.5)

                    print("✅ Reposición correcta")

                # Recapturar filas ya en página correcta
                filas = driver.find_elements(By.CSS_SELECTOR, "#tbData .FilaDatos")


            except:
                pass

            
            idx += 1

            
            
        # ===============================
        # 👉 PAGINACIÓN REAL (FUNCIONA EN TU WEB)
        # ===============================

        # ===============================s
        # 👉 PAGINACIÓN SEGURA Y VALIDADA
        # ===============================

        # ==========================================================
        # 👉 PAGINACIÓN REAL USANDO BOTÓN "Siguiente"
        # ==========================================================

        # ==========================================================
        # 👉 PAGINACIÓN POR CLASE ACTIVA (100% REAL)
        # ==========================================================

        


        
        
        # ==========================================================
        # 🧠 FORZAR POSTBACK SI LA PÁGINA TUVO SOLO VERDES
        # ==========================================================
        # ==========================================================
        # 🔥 FORZAR POSTBACK REAL SI TODA LA PÁGINA FUE VERDE
        # =========================================================     
        
        # 🔥 FIX CRÍTICO: FORZAR REPOSICIÓN ANTES DE PAGINAR
        print("🧠 Verificando posición real antes de paginar...")

        if pagina_actual > 1:
            ir_a_pagina(driver, pagina_actual)


        time.sleep(2)

        firma_pagina_actual = obtener_firma_pagina(driver)
        print(f"🔑 Firma real antes de paginar: {firma_pagina_actual}")
            
                
                                    
        firma_antes = obtener_firma_pagina(driver)

        exito = ir_a_siguiente_pagina_robusto(
            driver,
            pagina_actual
        )

        if not exito:
            print("🏁 No se pudo continuar — fin real")
            break

        # esperar que cambie la página real
        cambio_real = False
        for _ in range(10):

            time.sleep(1)

            firma_despues = obtener_firma_pagina(driver)

            if firma_despues != firma_antes:
                cambio_real = True
                break

        if not cambio_real:
            print("🏁 Página no cambió — fin de paginación")
            break

        if not cambio_real:
            print("🏁 Página no cambió — fin de paginación")
            break

        pagina_actual += 1
        print(f"➡ Página {pagina_actual} confirmada")



    # ==========================================================
    # 💾 GUARDAR EXCEL SIN ERRORES
    # ==========================================================

    if os.path.exists(ruta_excel):

        with pd.ExcelWriter(
            ruta_excel,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="replace"
        ) as writer:

            df.to_excel(writer, sheet_name=nombre_hoja, index=False)

    else:

        with pd.ExcelWriter(
            ruta_excel,
            engine="openpyxl",
            mode="w"
        ) as writer:

            df.to_excel(writer, sheet_name=nombre_hoja, index=False)

    print("\n✅ RESTRI.xlsx actualizado correctamente")
    
    print("\n✅ EXCEL FINAL GENERADO — TODAS LAS PÁGINAS")

    return restricciones_hechas
    
    
  
    
def extraer_tabla_a_excel_tuberias(driver, acuerdo_seleccionado):
    import os
    import time
    import pandas as pd
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait, Select
    from selenium.webdriver.support import expected_conditions as EC
    import re
    


    
    # ==========================================================
    #   🧠 DICCIONARIO DE MARCAS
    # ==========================================================
    MARCAS_DICC = [
        "SUPERTUBO", "GLOBALPLAST", "TUBOPLAST", "KRAH CHILE",
        "PREMIUM PLAST", "EUROTUBO", "PLASTIFORTE", "KUELAP",
        "JARVIS", "INSOMED", "ANDINACOLOR", "AYCER", "(H) TOOLS",
        "APOLO", "ECONORACKS", "TIGRE", "BUGUI (H)", "QM QUILMANS",
        "ANAX", "CERCOR", "A ALPETSA CALIDAD & GARANTIA",
        "NOVOPLAST", "PINTURAS UNO LA PINTURA DE LOS ESPECIALISTAS",
        "KOPLAST CONECTANDO EL PROGRESO", "DELSA", "CPP", "TEHMCO",
        "VENCEDOR", "INKATOOLS", "JAZCOR", "TACLLA", "TOOLBOX",
        "HILUX LA PINTURA TODO TERRENO", "MAXCER", "GAMAX",
        "BARITIA", "THN COLORS", "IMANINKI", "AVILSAA",
        "SERCONSLIMP", "THAILER", "ANKER PINTURAS INDUSTRIALES",
        "MAXTIC", "ECOCOLOR", "ARES PINTURAS",
        "INDIGO PINTURA Y LIMPIEZA DE ALTURA", "TOM",
        "CICLÓN LÍDER EN EL NORTE", "MARROK", "PILLKO",
        "VELSALIT", "ANDINACOAT TECNOLOGÍA EN RECUBRIMIENTOS",
        "AXMET", "INYECTOPLAST", "AMERICAN COLORS", "IS IVSACOR",
        "AUKI", "UNIVERSAL COLORS", "ISAVAL", "PLASTICA",
        "FAST", "PINTURAS PEGASO", "TEKNO", "MUCHICK",
        "JET", "U.S.COATINGS", "SUPER CAMPESINO", "TECPIPE",
        "INNOVA", "PALACIO COLORS", "QOSOFT", "WHUAYRA",
        "P PLASTITUBO PERU", "RAMISS", "MPLAST"
    ]
    
    def detectar_marca_por_diccionario(texto_ficha):
        texto = texto_ficha.upper()
        for marca_dic in MARCAS_DICC:
            if marca_dic in texto:
                return marca_dic
        return ""
    
    def esperar_loader_robusto(driver, timeout=15, timeout_total=30):
        inicio = time.time()

        while True:
            loaders = driver.find_elements(By.CSS_SELECTOR, ".loading, .spinner, .k-loading-mask")

            if not loaders:
                return True

            if time.time() - inicio > timeout_total:
                print("💀 LOADER CONGELADO DETECTADO")
                return False

            time.sleep(0.5)


    def recuperar_de_freeze(driver, pagina_actual):
        print("♻ RECUPERANDO SISTEMA POR FREEZE...")

        driver.refresh()

        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "#tbData"))
        )

        ir_a_pagina(driver, pagina_actual)

        print("✅ SISTEMA RECUPERADO")


    def obtener_pagina_activa(driver):
        try:
            return driver.find_element(
                By.CSS_SELECTOR,
                "#divPaginacion input.NavegaActivo"
            ).get_attribute("value")
        except:
            return "0"


    def ir_a_pagina(driver, numero_objetivo):

        print(f"🎯 Navegando inteligentemente a página {numero_objetivo}")

        try:
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.ID, "divPaginacion"))
            )

            pagina_actual_real = int(obtener_pagina_activa(driver))

            if pagina_actual_real == numero_objetivo:
                return True

            boton_directo = driver.find_elements(
                By.XPATH,
                f"//div[@id='divPaginacion']//input[@value='{numero_objetivo}']"
            )

            if boton_directo:
                driver.execute_script("arguments[0].click();", boton_directo[0])

                WebDriverWait(driver, 20).until(
                    lambda d: obtener_pagina_activa(d) == str(numero_objetivo)
                )

                return True

            botones_visibles = driver.find_elements(
                By.XPATH,
                "//div[@id='divPaginacion']//input[@type='button' and not(contains(@value,'Siguiente')) and not(contains(@value,'Anterior'))]"
            )

            numeros_visibles = []

            for b in botones_visibles:
                try:
                    val = int(b.get_attribute("value"))
                    numeros_visibles.append(val)
                except:
                    pass

            if not numeros_visibles:
                return False

            ancla = max([n for n in numeros_visibles if n < numero_objetivo], default=None)

            if not ancla:
                return False

            boton_ancla = driver.find_element(
                By.XPATH,
                f"//div[@id='divPaginacion']//input[@value='{ancla}']"
            )

            driver.execute_script("arguments[0].click();", boton_ancla)

            WebDriverWait(driver, 20).until(
                lambda d: obtener_pagina_activa(d) == str(ancla)
            )

            time.sleep(1)

            pagina_actual_real = ancla

            while pagina_actual_real < numero_objetivo:

                boton_siguiente = driver.find_element(
                    By.XPATH,
                    "//div[@id='divPaginacion']//input[contains(@value,'Siguiente')]"
                )

                driver.execute_script("arguments[0].click();", boton_siguiente)

                WebDriverWait(driver, 20).until(
                    lambda d: obtener_pagina_activa(d) == str(pagina_actual_real + 1)
                )

                pagina_actual_real += 1
                time.sleep(0.8)

            return True

        except Exception as e:
            print(f"💀 Error navegando a página {numero_objetivo}: {e}")
            return False


    def ir_a_siguiente_pagina_robusto(driver, pagina_actual):
        print("\n🔎 Intentando ir a la siguiente página...")
        return ir_a_pagina(driver, pagina_actual + 1)
        





            
            
    def obtener_firma_pagina(driver):
        filas = driver.find_elements(By.CSS_SELECTOR, "#tbData tr.FilaDatos")
        if not filas:
            return None
        return filas[0].find_element(By.CSS_SELECTOR, "td.primerTd").text.strip()

            
            
    




    

    print("\n📊 INICIANDO EXTRACCIÓN…")

    # Ajustar zoom
    try:
        driver.execute_script("document.body.style.zoom='75%'")
    except:
        pass

    ruta_excel = ruta_recurso("assets/EXTRAER.xlsx")
    
    # 🧠 Nombre de hoja = ACUERDO MARCO
    nombre_hoja = acuerdo_seleccionado.strip()[:31]  # Excel máximo 31 caracteres


    columnas = [
        "N°", "REQUERIMIENTO", "PROCEDIMIENTO", "PROFORMA",
        "FECHA DE EMISION", "FECHA LIMITE DE COTIZACION",
        "ENTIDAD", "RUC", "PRODUCTO", "FICHA PRODUCTO",
        "MARCA", "CODIGO", "CANTIDAD", "DIRECCION",
        "INICIO DE ENTREGA", "PLAZO MAXI", "FIN DE ENTREGA",
        "SUBTOTAL", "PDF" 
    ]


    # ==========================================================
    #   Cargar Excel
    # ==========================================================
    # ==========================================================
    # 📂 CARGAR O CREAR HOJA DEL ACUERDO
    # ==========================================================
    if os.path.exists(ruta_excel):
        try:
            df = pd.read_excel(ruta_excel, sheet_name=nombre_hoja)
            df = df.fillna("").astype(str).applymap(lambda x: x.strip())
            print(f"📂 Hoja '{nombre_hoja}' cargada.")
            
            # 🔒 ASEGURAR TODAS LAS COLUMNAS (ANTI KeyError)
            for col in columnas:
                if col not in df.columns:
                    df[col] = ""
        except:
            print(f"📄 Hoja '{nombre_hoja}' no existe. Se creará.")
            df = pd.DataFrame(columns=columnas)
    else:
        print("📂 Archivo no existe. Se creará nuevo.")
        df = pd.DataFrame(columns=columnas)
    # ==========================================================
    #   ✅ PASO 1 — CONTROL DE PAGINACIÓN
    # ==========================================================
    pagina_actual = 1

    while True:
        print(f"\n🧭 ===== PROCESANDO PÁGINA {pagina_actual} =====")

        # ==========================================================
        #   ✅ PASO 2 — CAPTURAR FILAS (DENTRO DEL WHILE)
        # ==========================================================
        filas = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "#tbData .FilaDatos"))
        )

        firma_pagina_actual = obtener_firma_pagina(driver)
        print(f"🔑 Firma página {pagina_actual}: {firma_pagina_actual}")

        
        

        # ==========================================================
        #   RECORRER FILAS (TU LÓGICA ORIGINAL, SIN CAMBIOS)
        # ==========================================================
        for idx in range(len(filas)):
            
            # 🔁 ASEGURAR QUE ESTAMOS EN LA PÁGINA CORRECTA
            if pagina_actual > 1 and idx > 0:
                firma_actual_en_vista = obtener_firma_pagina(driver)

                if firma_actual_en_vista != firma_pagina_actual:
                    print(f"🔄 Regreso detectado, volviendo a página {pagina_actual}")
                    ir_a_pagina(driver, pagina_actual)

                    WebDriverWait(driver, 15).until(
                        lambda d: obtener_firma_pagina(d) == firma_pagina_actual
                    )


            try:
                filas = WebDriverWait(driver, 10).until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, "#tbData .FilaDatos"))
                )
                fila = filas[idx]
            except:
                print("❌ No se pudo recapturar fila. Continuando…")
                continue

            print(f"\n----- FILA {idx+1} | PÁG {pagina_actual} -----")

            celdas = fila.find_elements(By.TAG_NAME, "td")

            # 1️⃣ Datos antes del modal
            try:
                requerimiento = celdas[0].text.strip()
                procedimiento = celdas[1].text.strip()
                proforma = celdas[4].text.strip()
                fecha_emision = celdas[5].text.strip()
                fecha_limite = celdas[8].text.strip()
                entidad = celdas[10].text.strip()
                ruc = celdas[11].text.strip()
            except:
                continue

            # 2️⃣ Click Cotizar
            try:
                boton = fila.find_element(By.CSS_SELECTOR, "button[title='Cotizar']")
                driver.execute_script("arguments[0].scrollIntoView(true);", boton)
                time.sleep(0.2)
                boton.click()
            except:
                continue

            # 3️⃣ Esperar modal
            try:
                WebDriverWait(driver, 10).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, "#divItem"))
                )
            except:
                continue

            # 4️⃣ Recorrer selects
            try:
                total_proforma = driver.find_element(By.ID, "lblTotal").text.strip()
                
                # 🧠 ESPERAR SELECTS O CONFIRMAR QUE NO EXISTEN
                selects = []

                try:
                    WebDriverWait(driver, 6).until(
                        EC.presence_of_element_located(
                            (By.CSS_SELECTOR, "#divItem select.proformasProducto")
                        )
                    )
                    selects = driver.find_elements(
                        By.CSS_SELECTOR, "#divItem select.proformasProducto"
                    )
                except:
                    selects = []


                # ==================================================
                # 🟡 CASO ESPECIAL: NO EXISTEN SELECTS (SOLO UNA OPCIÓN)
                # ==================================================
                if not selects:
                    print("⚠️ NO HAY SELECTS — SE DETECTA OPCIÓN ÚNICA")

                    try:
                        fila_producto = driver.find_element(
                            By.CSS_SELECTOR, "#divItem .flex-table.item"
                        )

                        celdas_prod = fila_producto.find_elements(By.CLASS_NAME, "flex-row")

                        producto = celdas_prod[0].text.strip()
                        ficha = celdas_prod[2].text.strip()
                        cantidad = celdas_prod[5].text.strip()
                        proforma_item = celdas_prod[6].text.strip()

                        # ----------------------------
                        # MARCA Y CÓDIGO (MISMA LÓGICA)
                        # ----------------------------
                        partes_ficha = ficha.split()
                        if len(partes_ficha) >= 2:
                            marca_default = partes_ficha[-2]
                            codigo = partes_ficha[-1]
                        else:
                            marca_default = ""
                            codigo = ""

                        marca_dic = detectar_marca_por_diccionario(ficha)
                        marca = marca_dic if marca_dic else marca_default

                        # ----------------------------
                        # PDF
                        # ----------------------------
                        try:
                            pdf_link = fila_producto.find_element(
                                By.CSS_SELECTOR, "a[href*='blob.core.windows.net']"
                            ).get_attribute("href").strip()
                        except:
                            pdf_link = ""

                        # ----------------------------
                        # SUBTOTAL (ENTREGAS — ÚNICA OPCIÓN)
                        # ----------------------------
                        try:
                            fila_entrega = driver.find_element(
                                By.CSS_SELECTOR,
                                "#divEntregas .flex-table.item[data-id]"
                            )
                            celdas_entrega = fila_entrega.find_elements(By.CLASS_NAME, "flex-row")
                            subtotal = celdas_entrega[-1].text.strip()
                        except:
                            subtotal = ""

                        # ----------------------------
                        # DIRECCIÓN / FECHAS (CORREGIDO)
                        # ----------------------------
                        cab = driver.find_element(
                            By.CSS_SELECTOR,
                            "#divEntregas .flex-table.item[data-id]"
                        ).find_elements(By.CLASS_NAME, "flex-row")

                        direccion = cab[1].text.strip()
                        inicio = cab[2].text.strip()
                        fin = cab[4].text.strip()

                        # 🔥 PLAZO MAXI (INPUT O TEXTO)
                        plazo = ""

                        try:
                            # 1️⃣ Intentar desde input (primer caso)
                            plazo_input = driver.find_element(
                                By.CSS_SELECTOR,
                                "input.entregas_n_plazo"
                            )
                            plazo = plazo_input.get_attribute("value").strip()

                        except:
                            try:
                                # 2️⃣ Fallback: desde columna "Plazo máximo"
                                cab = driver.find_element(
                                    By.CSS_SELECTOR,
                                    "#divEntregas .flex-table.item[data-id]"
                                ).find_elements(By.CLASS_NAME, "flex-row")

                                plazo = cab[3].text.strip()  # 👈 AQUÍ ESTÁ EL 10
                            except:
                                plazo = ""



                        # ----------------------------
                        # GUARDAR (MISMO FORMATO)
                        # ----------------------------
                        clave = (requerimiento, proforma_item, codigo, ruc, ficha)

                        existe = df[
                            (df["REQUERIMIENTO"] == clave[0]) &
                            (df["PROFORMA"] == clave[1]) &
                            (df["CODIGO"] == clave[2]) &
                            (df["RUC"] == clave[3]) &
                            (df["FICHA PRODUCTO"] == clave[4])
                        ]

                        if existe.empty:
                            df = pd.concat([df, pd.DataFrame([{
                                "N°": len(df) + 1,
                                "REQUERIMIENTO": requerimiento,
                                "PROCEDIMIENTO": procedimiento,
                                "PROFORMA": proforma_item,
                                "FECHA DE EMISION": fecha_emision,
                                "FECHA LIMITE DE COTIZACION": fecha_limite,
                                "ENTIDAD": entidad,
                                "RUC": ruc,
                                "PRODUCTO": producto,
                                "FICHA PRODUCTO": ficha,
                                "MARCA": marca,
                                "CODIGO": codigo,
                                "CANTIDAD": cantidad,
                                "DIRECCION": direccion,
                                "INICIO DE ENTREGA": inicio,
                                "PLAZO MAXI": plazo,
                                "FIN DE ENTREGA": fin,
                                "SUBTOTAL": subtotal,
                                "PDF": pdf_link
                            }])], ignore_index=True)

                            print("✅ GUARDADO (OPCIÓN ÚNICA SIN SELECT)")
                            
                            # ----------------------------
                            # 🔒 CERRAR MODAL Y CONTINUAR
                            # ----------------------------
                            # 🔒 CERRAR MODAL AL FINAL DE CADA FILA
                            # ==================================================
                            # 🔒 CIERRE ROBUSTO REAL DEL MODAL (VERSIÓN BUENA)
                            # ==================================================
                            try:
                                print("🔒 Intentando cerrar modal...")

                                # 1️⃣ Esperar que botón exista (NO clickable)
                                WebDriverWait(driver, 10).until(
                                    EC.presence_of_element_located((By.ID, "btnCerraPopupCotizacion"))
                                )

                                # 2️⃣ Click por JavaScript (NO .click())
                                driver.execute_script("""
                                    var btn = document.getElementById("btnCerraPopupCotizacion");
                                    if(btn){ btn.click(); }
                                """)

                                # 3️⃣ Esperar que el contenedor del modal desaparezca
                                WebDriverWait(driver, 20).until(
                                    EC.invisibility_of_element_located((By.CSS_SELECTOR, "#divItem"))
                                )

                                # 4️⃣ Esperar que la tabla vuelva activa
                                WebDriverWait(driver, 20).until(
                                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, "#tbData .FilaDatos"))
                                )

                                time.sleep(1)

                                print("✅ Modal cerrado 100% correctamente")

                            except Exception as e:
                                print("💀 Error cerrando modal:", e)

                            continue  # 🔴 IR A LA SIGUIENTE FILA


                    except Exception as e:
                        print("❌ ERROR CASO SIN SELECT:", e)

                    continue  # 🔴 IMPORTANTE: saltar lógica de selects


                for s_idx in range(len(selects)):
                    selects = driver.find_elements(
                        By.CSS_SELECTOR, "#divItem select.proformasProducto"
                    )
                    select = Select(selects[s_idx])

                    for o_idx in range(1, len(select.options)):
                        select = Select(
                            driver.find_elements(
                                By.CSS_SELECTOR, "#divItem select.proformasProducto"
                            )[s_idx]
                        )

                        opcion = select.options[o_idx]
                        ficha = opcion.text.strip()
                        
                        print(f"\n🧾 FICHA PRODUCTO DETECTADA: {ficha}")

                        # ==================================================
                        # EXTRAER MARCA Y CODIGO DESDE FICHA PRODUCTO
                        # (SIEMPRE LAS ÚLTIMAS 2 PALABRAS)
                        # ==================================================
                        partes_ficha = ficha.split()

                        # 🔹 Marca y código por lógica original
                        if len(partes_ficha) >= 2:
                            marca_default = partes_ficha[-2]
                            codigo = partes_ficha[-1]
                        else:
                            marca_default = ""
                            codigo = ficha
                            
                        print(f"🏷️ CÓDIGO EXTRAÍDO: {codigo}")


                        # 🔍 Buscar marca en diccionario
                        marca_dic = detectar_marca_por_diccionario(ficha)
                        
                        if marca_dic:
                                print(f"🔍 MARCA ENCONTRADA EN DICCIONARIO: {marca_dic}")
                        else:
                            print("⚠️ NO SE ENCONTRÓ MARCA EN DICCIONARIO")


                        # ✅ Prioridad: diccionario > lógica original
                        marca = marca_dic if marca_dic else marca_default
                        
                        origen = "DICCIONARIO" if marca_dic else "LÓGICA ORIGINAL"
                        print(f"✅ MARCA FINAL USADA: {marca}  (ORIGEN: {origen})")





                        valor = opcion.get_attribute("value")
                        if not valor or "^" not in valor:
                            continue

                        ##codigo = valor.split("^")[1]

                        select.select_by_index(o_idx)
                        time.sleep(0.6)
                        
                        # ==========================
                        # 📄 EXTRAER PDF DE ESTA OPCIÓN
                        # ==========================
                        pdf_link = ""

                        try:
                            fila_producto = driver.find_elements(
                                By.CSS_SELECTOR, "#divItem .flex-table.item"
                            )[s_idx]

                            pdf_a = fila_producto.find_element(
                                By.CSS_SELECTOR, "a[href*='blob.core.windows.net']"
                            )

                            pdf_link = pdf_a.get_attribute("href").strip()

                        except:
                            pdf_link = ""

                        print(f"📎 PDF DETECTADO: {pdf_link}")

                        
                        # ==================================================
                        # EXTRAER SUBTOTAL DEL PRODUCTO (MISMA FILA)
                        # ==================================================

                        try:
                            filas_productos_entrega = driver.find_elements(
                                By.CSS_SELECTOR,
                                "#divEntregas .table-container[style*='margin-left: 3%'] .flex-table.item"
                            )

                            fila_subtotal = filas_productos_entrega[s_idx]
                            celdas_sub = fila_subtotal.find_elements(By.CLASS_NAME, "flex-row")

                            # 👉 ÚLTIMA COLUMNA = SUBTOTAL
                            subtotal = celdas_sub[-1].text.strip()

                            print(f"💰 SUBTOTAL OPCIÓN {s_idx + 1}: {subtotal}")

                        except Exception as e:
                            subtotal = ""
                            print("⚠️ No se pudo extraer subtotal por opción:", e)





                        fila_producto = driver.find_elements(
                            By.CSS_SELECTOR, "#divItem .flex-table.item"
                        )[s_idx]

                        celdas_prod = fila_producto.find_elements(By.CLASS_NAME, "flex-row")

                        producto = celdas_prod[0].text.strip()
                        cantidad = celdas_prod[5].text.strip()
                        proforma_item = celdas_prod[6].text.strip()



                        cab = driver.find_element(
                            By.CSS_SELECTOR,
                            "#divEntregas .flex-table.item[data-id]"
                        ).find_elements(By.CLASS_NAME, "flex-row")

                        direccion = cab[1].text.strip()
                        inicio = cab[2].text.strip()
                        plazo = cab[3].text.strip()
                        fin = cab[4].text.strip()
                        


                        clave = (requerimiento, proforma_item, codigo, ruc, ficha)

                        existe = df[
                            (df["REQUERIMIENTO"] == clave[0]) &
                            (df["PROFORMA"] == clave[1]) &
                            (df["CODIGO"] == clave[2]) &
                            (df["RUC"] == clave[3]) &
                            (df["FICHA PRODUCTO"] == clave[4])
                        ]

                        if existe.empty:
                            print("💾 REGISTRO NUEVO → SE GUARDARÁ EN EXCEL")

                            df = pd.concat([df, pd.DataFrame([{
                                "N°": len(df) + 1,
                                "REQUERIMIENTO": requerimiento,
                                "PROCEDIMIENTO": procedimiento,
                                "PROFORMA": proforma_item,
                                "FECHA DE EMISION": fecha_emision,
                                "FECHA LIMITE DE COTIZACION": fecha_limite,
                                "ENTIDAD": entidad,
                                "RUC": ruc,
                                "PRODUCTO": producto,
                                "FICHA PRODUCTO": ficha,
                                "MARCA": marca,
                                "CODIGO": codigo,
                                "CANTIDAD": cantidad,
                                "DIRECCION": direccion,
                                "INICIO DE ENTREGA": inicio,
                                "PLAZO MAXI": plazo,
                                "FIN DE ENTREGA": fin,
                                "SUBTOTAL": subtotal,
                                "PDF": pdf_link 
                                
                            
                            }])], ignore_index=True)
                            
                            print(f"📥 GUARDADO | REQ: {requerimiento} | PROFORMA: {proforma_item}")

            except Exception as e:
                print("❌ Error dentro del modal (tuberías):", e)
            finally:

                # ==================================================
                # 🔒 CIERRE ROBUSTO REAL DEL MODAL (VERSIÓN BUENA)
                # ==================================================
                try:
                    print("🔒 Intentando cerrar modal...")

                    # 1️⃣ Esperar que botón exista (NO clickable)
                    WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "btnCerraPopupCotizacion"))
                    )

                    # 2️⃣ Click por JavaScript (NO .click())
                    driver.execute_script("""
                        var btn = document.getElementById("btnCerraPopupCotizacion");
                        if(btn){ btn.click(); }
                    """)

                    # 3️⃣ Esperar que el contenedor del modal desaparezca
                    WebDriverWait(driver, 20).until(
                        EC.invisibility_of_element_located((By.CSS_SELECTOR, "#divItem"))
                    )

                    # 4️⃣ Esperar que la tabla vuelva activa
                    WebDriverWait(driver, 20).until(
                        EC.presence_of_all_elements_located((By.CSS_SELECTOR, "#tbData .FilaDatos"))
                    )

                    time.sleep(1)

                    print("✅ Modal cerrado 100% correctamente")

                except Exception as e:
                    print("💀 Error cerrando modal:", e)

        # ===============================
        # 👉 PAGINACIÓN REAL (ROBUSTA)
        # ===============================

        # ======================================================
        # 🧠 FORZAR POSICIÓN REAL ANTES DE PAGINAR
        # ======================================================

        print("🧠 Verificando posición real antes de paginar...")

        if pagina_actual > 1:
            ir_a_pagina(driver, pagina_actual)

        time.sleep(1)
        
        
        # 🛡️ ESPERAR TABLA ESTABLE ANTES DE PAGINAR
        WebDriverWait(driver, 20).until(
            EC.presence_of_all_elements_located(
                (By.CSS_SELECTOR, "#tbData .FilaDatos")
            )
        )
        time.sleep(1)

        # ======================================================
        # 🧠 PAGINACIÓN ROBUSTA
        # ======================================================

        exito = ir_a_siguiente_pagina_robusto(driver, pagina_actual)

        if not exito:
            print("🏁 No se pudo continuar — fin real")
            break

        ok_loader = esperar_loader_robusto(driver, 15, 25)

        if not ok_loader:
            recuperar_de_freeze(driver, pagina_actual)

        pagina_actual += 1
        print(f"➡ Página {pagina_actual} confirmada")

        print("⏳ Esperando carga completa de la tabla...")

        if not esperar_loader_robusto(driver, 20, 40):
            print("💀 Freeze detectado tras paginar — recuperando")
            recuperar_de_freeze(driver, pagina_actual)

        time.sleep(1)



    # ==========================================================
    #   GUARDAR EXCEL FINAL
    # ==========================================================
    with pd.ExcelWriter(
        ruta_excel,
        engine="openpyxl",
        mode="a" if os.path.exists(ruta_excel) else "w",
        if_sheet_exists="replace"
    ) as writer:
        df.to_excel(writer, sheet_name=nombre_hoja, index=False)
    print("\n✅ EXCEL FINAL GENERADO — TODAS LAS PÁGINAS")
    
    
    
import pandas as pd
import matplotlib.pyplot as plt

def analizar_excel_ventas(ruta_excel):
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    # ==========================
    # 1️⃣ CARGAR DATOS
    # ==========================
    df = pd.read_excel(ruta_excel)

    df["SUBTOTAL"] = pd.to_numeric(df["SUBTOTAL"], errors="coerce").fillna(0)
    df["CANTIDAD"] = pd.to_numeric(df["CANTIDAD"], errors="coerce").fillna(0)
    df["FECHA DE EMISION"] = pd.to_datetime(df["FECHA DE EMISION"], errors="coerce")

    # ==========================
    # 2️⃣ CLASIFICACIÓN DE PRODUCTOS
    # ==========================
    def clasificar_producto(texto):
        texto = str(texto).upper()
        if any(x in texto for x in ["TACHO", "BOLSA", "DETERGENTE", "LIMPIADOR"]):
            return "LIMPIEZA"
        if any(x in texto for x in ["TUBO", "PINTURA", "CERAM", "SANIT"]):
            return "FERRETERÍA"
        if any(x in texto for x in ["CONTENEDOR", "RECOLECTOR", "PLÁSTICO"]):
            return "PLÁSTICOS"
        if any(x in texto for x in ["PAPEL", "OFICINA", "TONER"]):
            return "OFICINA"
        return "OTROS"

    df["CATEGORIA"] = df["PRODUCTO"].apply(clasificar_producto)

    # ==========================
    # 3️⃣ CREAR NUEVO EXCEL
    # ==========================
    ruta_dashboard = ruta_excel.replace(".xlsx", "_DASHBOARD.xlsx")

    with pd.ExcelWriter(ruta_dashboard, engine="openpyxl") as writer:

        # BASE
        df.to_excel(writer, sheet_name="BASE", index=False)

        # KPIS
        kpis = pd.DataFrame({
            "INDICADOR": [
                "Total Ventas (S/)",
                "Total Cantidad",
                "N° Entidades",
                "N° Productos",
                "Ticket Promedio"
            ],
            "VALOR": [
                df["SUBTOTAL"].sum(),
                df["CANTIDAD"].sum(),
                df["ENTIDAD"].nunique(),
                df["PRODUCTO"].nunique(),
                round(df["SUBTOTAL"].sum() / max(len(df), 1), 2)
            ]
        })
        kpis.to_excel(writer, sheet_name="KPIS", index=False)

        # ENTIDADES
        df.groupby("ENTIDAD", as_index=False)["SUBTOTAL"].sum() \
          .sort_values("SUBTOTAL", ascending=False) \
          .to_excel(writer, sheet_name="ENTIDADES", index=False)

        # PRODUCTOS
        df.groupby("PRODUCTO", as_index=False)["SUBTOTAL"].sum() \
          .sort_values("SUBTOTAL", ascending=False) \
          .to_excel(writer, sheet_name="PRODUCTOS", index=False)

        # MARCAS
        df.groupby("MARCA", as_index=False)["SUBTOTAL"].sum() \
          .sort_values("SUBTOTAL", ascending=False) \
          .to_excel(writer, sheet_name="MARCAS", index=False)

        # CATEGORÍAS
        df.groupby("CATEGORIA", as_index=False)["SUBTOTAL"].sum() \
          .sort_values("SUBTOTAL", ascending=False) \
          .to_excel(writer, sheet_name="CATEGORIAS", index=False)

        # TIEMPO
        df["MES"] = df["FECHA DE EMISION"].dt.to_period("M").astype(str)
        df.groupby("MES", as_index=False)["SUBTOTAL"].sum() \
          .to_excel(writer, sheet_name="TIEMPO", index=False)

    # ==========================
    # 4️⃣ FORMATO PROFESIONAL
    # ==========================
    wb = load_workbook(ruta_dashboard)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        ws.auto_filter.ref = ws.dimensions
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = 22
            col[0].font = Font(bold=True)

    wb.save(ruta_dashboard)

    print("✅ Dashboard Excel creado correctamente:")
    print(f"📊 {ruta_dashboard}")




# ==========================================================
# 🔹 CARGAR PRODUCTOS AUTOMÁTICAMENTE + CARACTERÍSTICAS
# ==========================================================
def cargar_productos(driver, wait, combinaciones, tipo_producto, datos_doc, codigos_finales=None):
    
    """Carga imágenes, fichas, precios y luego agrega características por variante."""
    print("📦 Iniciando carga de productos...")
    
    caracteristicas = datos_doc.get("caracteristicas", {})

    # Buscar cualquier clave que contenga "CODIGO" y "BASE" ignorando mayúsculas y acentos
    codigo_base = None
    for key, value in caracteristicas.items():
        key_norm = key.upper().replace("Ó", "O").replace(" ", "_")
        if "CODIGO_BASE" in key_norm:
            codigo_base = value.strip()
            break

    if not codigo_base:
        print("❌ ERROR: No se encontró CODIGO_BASE en 'datos_doc'")
        return


    carpeta_base = rf"D:\PCOMPRAS\{tipo_producto}"  # carpeta que el usuario seleccionó


    carpeta_imagenes = os.path.join(carpeta_base, codigo_base.upper(), "IMAGENES")
    carpeta_fichas = os.path.join(carpeta_base, codigo_base.upper(), "FICHAS")

    
    




    def buscar_archivo_por_codigo(carpeta, codigo, extensiones):
        """Busca un archivo en la carpeta cuyo nombre coincida exactamente con el código final."""
        import os

        if not os.path.exists(carpeta):
            print(f"❌ Carpeta no encontrada: {carpeta}")
            return None

        for ext in extensiones:
            archivo = f"{codigo}{ext}"
            ruta = os.path.join(carpeta, archivo)
            if os.path.exists(ruta) or os.path.exists(ruta.lower()):
                print(f"✅ Archivo encontrado para '{codigo}': {ruta}")
                return ruta

        print(f"❌ No se encontró archivo para '{codigo}' en {carpeta}")
        return None





    for i in range(len(codigos_finales)):
        color_detectado, unidad, precio_general = combinaciones[i]
        codigo_final = codigos_finales[i]

        print(f"\n🆔 Código cargando: {codigo_final} | 🎨 Color: {color_detectado} | 🧾 Unidad: {unidad} | 💰 Precio: {precio_general}")


        print(f"\n🎨 Procesando color: {color_detectado} | 💰 Precio único: {precio_general}")

        ruta_imagen = buscar_archivo_por_codigo(carpeta_imagenes, codigo_final, [".JPG", ".JPEG", ".PNG"])
        ruta_ficha = buscar_archivo_por_codigo(carpeta_fichas, codigo_final, [".PDF"])


        if not ruta_imagen:
            print(f"⚠️ No se encontró imagen para el código {codigo_final}")
            continue
        if not ruta_ficha:
            print(f"⚠️ No se encontró ficha técnica PDF para el código {codigo_final}")
            continue

        try:
            print(f"🆕 Iniciando carga para ficha {codigo_final} en formulario abierto.")


            # Ahora la carga se hace igual que antes
            upload_img = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@type='file' and @id='upload']")))
            upload_img.send_keys(ruta_imagen)
            print(f"🖼️ Imagen cargada correctamente para {codigo_final}")

            upload_pdf = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@type='file' and @id='uploadAdj']")))
            upload_pdf.send_keys(ruta_ficha)
            print(f"📄 Ficha técnica cargada correctamente para {codigo_final}")

            campo_precio = wait.until(EC.presence_of_element_located((By.ID, "C_PrecioRef")))
            campo_precio.clear()
            campo_precio.send_keys(precio_general)
            print(f"💰 Precio ingresado: {precio_general}")
            time.sleep(1)

            # Guardar producto
            try:
                btn_guardar = wait.until(EC.element_to_be_clickable((By.ID, "btnGuardar")))
                driver.execute_script("arguments[0].click();", btn_guardar)
                print(f"💾 Click en 'Guardar' para producto '{color_detectado}' realizado.")

                try:
                    form = driver.find_element(By.XPATH, "//form[contains(@id,'FormProducto')]")
                    driver.execute_script("arguments[0].submit();", form)
                except:
                    pass
            except Exception as e:
                print(f"⚠️ No se pudo hacer clic en 'Guardar': {e}")
                continue

            # Confirmar guardado
            time.sleep(3)
            print(f"✅ Producto '{color_detectado}' guardado correctamente.")

            # 🔹 AGREGAR CARACTERÍSTICAS DESPUÉS DE GUARDAR (bloqueante)
            try:
                print(f"🧩 Iniciando características para '{color_detectado}'...")
                agregar_caracteristicas(driver, wait, datos_doc["caracteristicas"], [color_detectado], codigo_final)
                


                print(f"✅ Características completadas para '{color_detectado}'.")
            except Exception as e:
                print(f"⚠️ Error al agregar características de '{color_detectado}': {e}")

            # 🔁 Preparar siguiente producto automáticamente
            try:
                print("↩️ Regresando al listado principal de productos...")
                btn_retornar = wait.until(EC.element_to_be_clickable((By.ID, "btnRegresarIndex")))
                driver.execute_script("arguments[0].scrollIntoView(true);", btn_retornar)
                driver.execute_script("arguments[0].click();", btn_retornar)

                # Esperar que cargue el listado
                btn_crear_nuevo = WebDriverWait(driver, 15).until(
                    EC.element_to_be_clickable((By.ID, "btnNuevoProducto"))
                )
                print("✅ Retorno al listado confirmado.")

                # Crear nuevo producto inmediatamente
                driver.execute_script("arguments[0].click();", btn_crear_nuevo)
                print("🆕 Click en 'Crear Producto' realizado para el siguiente color.")

                # Esperar a que el formulario se recargue completamente
                WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.ID, "C_PrecioRef"))
                )
                print("✅ Formulario nuevo listo, comenzando siguiente carga...\n")

                # 🔄 Continuar automáticamente al siguiente color
                continue

            except Exception as e:
                print(f"⚠️ Error al retornar o preparar siguiente producto: {e}")
                print("🔁 Intentando continuar con el siguiente color igualmente...")
                continue

        except Exception as e:
            print(f"⚠️ Error cargando producto '{color_detectado}': {e}")
            continue

    print("🏁 Carga masiva completada correctamente.")
