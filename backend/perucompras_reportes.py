"""
Helbot - perucompras_reportes.py
---------------------------------
Genera los mismos 3 reportes que el script standalone original
(extractor_proformas.py + filtrar_marcas_proformas.py + zt.py), pero:
  - Las listas de marcas (restringidas por semáforo, prohibidas entre
    S/500-1000, excepciones, marcas objetivo) ya NO están hardcodeadas
    aquí — se leen de la tabla perucompras_marcas_config, editable
    desde el frontend.
  - Cada fila "restringida" (semáforo o monto mínimo) se guarda también
    en MySQL (tabla perucompras_restringidos), no solo en el Excel.
  - Los 3 archivos Excel (historial del día, marcas, restringidos) se
    siguen generando igual que antes, en una carpeta configurable.
"""
import os
import logging
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from db import get_conn

logger = logging.getLogger("helbot.perucompras_reportes")

REPORTES_DIR = os.getenv(
    "PERUCOMPRAS_REPORTES_DIR",
    os.path.join(os.path.expanduser("~"), "Documents", "PeruComprasBot"),
)

PARES_COLORES = [
    ("FFF200", "FFF9B1"), ("4F81BD", "DCE6F1"), ("70AD47", "E2F0D9"),
    ("C0504D", "F2DCDB"), ("8064A2", "E4DFEC"), ("F79646", "FDE9D9"),
    ("00B0F0", "D9F2FF"), ("FF66CC", "FFD9F2"),
]

# ==========================================================
# ORDEN DE COLUMNAS PARA LOS EXCEL EXPORTADOS
# ----------------------------------------------------------
# Pedido por el usuario de extracción. En MySQL las columnas NO cambian,
# esto solo reordena lo que se escribe en el .xlsx.
# ==========================================================
ORDEN_COLUMNAS_EXCEL = [
    "N°", "FECHA_GUARDADO", "REQUERIMIENTO", "PROFORMA", "N_PROFORMA_ID",
    "N_ENTIDAD_SEMAFORO", "COLOR_SEMAFORO", "PROCEDIMIENTO", "FECHA_EMISION",
    "FECHA_LIMITE_COTIZACION", "ENTIDAD", "RUC", "PRODUCTO", "FICHA_PRODUCTO",
    "MARCA", "CODIGO_UNICO", "CANTIDAD", "PRECIO_UNITARIO_BASE",
    "PRECIO_OFERTADO", "MONEDA", "DIRECCION_ENTREGA", "DEPARTAMENTO",
    "PROVINCIA", "DISTRITO", "FECHA_INICIO_ENTREGA", "FECHA_FIN_ENTREGA",
    "PLAZO_DIAS", "SUBTOTAL", "COSTO_PRODUCTOS", "COSTO_ENVIO", "IGV",
    "PDF_PRODUCTO", "PDF_REQUERIMIENTO", "IMAGEN_PRODUCTO",
    # No venían en la lista pedida — se dejan al final para no perder el
    # dato. Si no se quieren en el Excel, se borran de aquí.
    "ESTADO", "UID_PERUCOMPRAS", "DETALLE_ENTREGA_ID",
]

# Mismo orden pero con los nombres reales de columna en MySQL (minúsculas),
# usado solo en generar_excel_historial_acumulado (que lee con SELECT *).
ORDEN_COLUMNAS_DB = [
    "id", "fecha_guardado", "requerimiento", "proforma", "n_proforma_id",
    "n_entidad_semaforo", "color_semaforo", "procedimiento", "fecha_emision",
    "fecha_limite_cotizacion", "entidad", "ruc", "producto", "ficha_producto",
    "marca", "codigo_unico", "cantidad", "precio_unitario_base",
    "precio_ofertado", "moneda", "direccion_entrega", "departamento",
    "provincia", "distrito", "fecha_inicio_entrega", "fecha_fin_entrega",
    "plazo_dias", "subtotal", "costo_productos", "costo_envio", "igv",
    "pdf_producto", "pdf_requerimiento", "imagen_producto",
    "estado", "uid_perucompras", "detalle_entrega_id", "catalogo",
    "ficha_hash", "prod_idx", "pf_idx", "det_idx",
]


def _reordenar_columnas(df: pd.DataFrame, orden: list[str]) -> pd.DataFrame:
    """Reordena columnas de df según `orden`. Lo que esté en `orden` pero
    no exista en df se ignora (no truena). Lo que df tenga y no esté en
    `orden` se agrega al final, para nunca perder datos por descuadre
    entre esta lista y la tabla real."""
    cols_presentes = [c for c in orden if c in df.columns]
    cols_extra = [c for c in df.columns if c not in orden]
    return df[cols_presentes + cols_extra]

def obtener_marcas_config(uid: str = "") -> dict[str, set[str]]:
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            if uid:
                cur.execute("SELECT lista, marca FROM perucompras_marcas_config WHERE uid_perucompras = %s", (uid,))
            else:
                cur.execute("SELECT lista, marca FROM perucompras_marcas_config")
            filas = cur.fetchall()
    finally:
        conn.close()
    resultado: dict[str, set[str]] = {
        "restringida_semaforo": set(),
        "prohibida_500_1000": set(),
        "excepcion_menor_500": set(),
        "objetivo": set(),
    }
    for f in filas:
        if f["lista"] in resultado:
            resultado[f["lista"]].add(f["marca"].strip().upper())
    return resultado


def clasificar_restringidos(registros_con_id: list[dict], marcas_config: dict[str, set[str]]) -> list[dict]:
    """Misma lógica que zt.py, fila por fila, sobre registros YA
    insertados en MySQL (cada uno trae su '_id' real)."""
    restringidas = marcas_config["restringida_semaforo"]
    prohibidas_500_1000 = marcas_config["prohibida_500_1000"]
    excepcion_menor_500 = marcas_config["excepcion_menor_500"]

    resultado = []
    for r in registros_con_id:
        procedimiento = str(r.get("PROCEDIMIENTO", "")).strip().upper()
        if procedimiento != "ORDINARIA - INDIVIDUAL":
            continue

        # Si la proforma YA tiene un estado resuelto en Perú Compras
        # (ya fue restringida, cotizada, o quedó desierta), no hay
        # ninguna acción pendiente real sobre ella — no debe volver a
        # aparecer como candidata en cada nueva extracción.
        estado_real = str(r.get("ESTADO") or "").strip().upper()
        if estado_real and estado_real != "PENDIENTE":
            continue

        marca = str(r.get("MARCA", "")).strip().upper()
        semaforo = str(r.get("COLOR_SEMAFORO", "")).strip().upper()
        subtotal = float(r.get("SUBTOTAL") or 0)

        if marca in restringidas and semaforo == "ROJO":
            resultado.append({"extraccion_id": r["_id"], "motivo": "semaforo", "marca": marca, "subtotal": subtotal})
            continue

        if subtotal > 1000:
            continue

        es_menor_500 = subtotal < 500 and marca not in excepcion_menor_500
        es_500_1000 = 500 <= subtotal <= 1000 and marca not in prohibidas_500_1000

        if es_menor_500 or es_500_1000:
            resultado.append({"extraccion_id": r["_id"], "motivo": "monto_minimo", "marca": marca, "subtotal": subtotal})

    return resultado


def guardar_restringidos_mysql(run_id: int | None, catalogo: str, filas: list[dict]):
    if not filas:
        return
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            for f in filas:
                cur.execute(
                    """
                    INSERT IGNORE INTO perucompras_restringidos
                        (run_id, extraccion_id, catalogo, motivo, marca, subtotal, estado)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """,
                    (run_id, f["extraccion_id"], catalogo, f["motivo"], f["marca"], f["subtotal"], "pendiente"),
                )
        conn.commit()
    finally:
        conn.close()

def clasificar_marcas_objetivo(registros_con_id: list[dict], marcas_config: dict[str, set[str]]) -> list[dict]:
    """Filas cuya marca está en la lista 'objetivo' — mismo criterio que
    usa generar_excel_marcas para armar el Excel, pero ahora también
    persistido en MySQL para que el equipo pueda rellenar proveedor y
    precio por cada una desde la interfaz."""
    objetivo = marcas_config["objetivo"]
    resultado = []
    for r in registros_con_id:
        marca = str(r.get("MARCA", "")).strip().upper()
        if marca in objetivo:
            resultado.append({"extraccion_id": r["_id"], "marca": marca})
    return resultado


def guardar_marcas_objetivo_mysql(run_id: int | None, catalogo: str, filas: list[dict]):
    if not filas:
        return
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            for f in filas:
                cur.execute(
                    """
                    INSERT IGNORE INTO perucompras_marcas_objetivo
                        (run_id, extraccion_id, catalogo, marca)
                    VALUES (%s, %s, %s, %s)
                    """,
                    (run_id, f["extraccion_id"], catalogo, f["marca"]),
                )
        conn.commit()
    finally:
        conn.close()


def generar_excel_historial(catalogos_registros: dict[str, list[dict]]) -> str | None:
    os.makedirs(REPORTES_DIR, exist_ok=True)
    fecha_hoy = datetime.now().strftime("%Y%m%d_%H%M%S")
    ruta = os.path.join(REPORTES_DIR, f"proformas_{fecha_hoy}.xlsx")

    if not any(catalogos_registros.values()):
        logger.info("generar_excel_historial: sin registros en ningún catálogo, se omite")
        return None

    with pd.ExcelWriter(ruta, engine="openpyxl") as writer:
        for catalogo, registros in catalogos_registros.items():
            if not registros:
                continue
            df = pd.DataFrame(registros).drop(columns=["_id"], errors="ignore")
            df = _reordenar_columnas(df, ORDEN_COLUMNAS_EXCEL)
            df.to_excel(writer, sheet_name=catalogo[:31], index=False)

    logger.info(f"generar_excel_historial: guardado en {ruta}")
    return ruta


def generar_excel_historial_acumulado(uid: str) -> str | None:
    """A diferencia de generar_excel_historial (que solo trae lo de la
    corrida actual), esto trae TODO lo que existe en MySQL para este
    uid desde el principio — cada extracción regenera este archivo
    completo desde cero a partir de la tabla, así siempre queda al día
    sin necesidad de ir "pegando" filas a un Excel viejo."""
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT DISTINCT catalogo FROM perucompras_extraccion WHERE uid_perucompras = %s ORDER BY catalogo",
                (uid,),
            )
            catalogos = [f["catalogo"] for f in cur.fetchall()]

            if not catalogos:
                return None

            os.makedirs(REPORTES_DIR, exist_ok=True)
            fecha_hoy = datetime.now().strftime("%Y%m%d_%H%M%S")
            ruta = os.path.join(REPORTES_DIR, f"historial_acumulado_{uid}_{fecha_hoy}.xlsx")

            hubo_contenido = False
            with pd.ExcelWriter(ruta, engine="openpyxl") as writer:
                for catalogo in catalogos:
                    cur.execute(
                        "SELECT * FROM perucompras_extraccion WHERE uid_perucompras = %s AND catalogo = %s ORDER BY fecha_guardado DESC",
                        (uid, catalogo),
                    )
                    filas = cur.fetchall()
                    if not filas:
                        continue
                    df = pd.DataFrame(filas)
                    df = _reordenar_columnas(df, ORDEN_COLUMNAS_DB)
                    df.to_excel(writer, sheet_name=catalogo[:31], index=False)
                    hubo_contenido = True

            if not hubo_contenido:
                try:
                    os.remove(ruta)
                except OSError:
                    pass
                return None

            logger.info(f"generar_excel_historial_acumulado: guardado en {ruta} ({len(catalogos)} catálogos, uid={uid})")
            return ruta
    finally:
        conn.close()



def generar_excel_marcas(ruta_historial: str, marcas_objetivo: set[str]) -> str | None:
    if not ruta_historial or not os.path.exists(ruta_historial) or not marcas_objetivo:
        return None

    nombre_salida = ruta_historial.replace(".xlsx", "_MARCAS.xlsx")
    xls = pd.ExcelFile(ruta_historial)
    hubo_contenido = False

    with pd.ExcelWriter(nombre_salida, engine="openpyxl") as writer:
        for hoja in xls.sheet_names:
            df = pd.read_excel(ruta_historial, sheet_name=hoja)
            if df.empty or "MARCA" not in df.columns or "PROFORMA" not in df.columns:
                continue
            df["MARCA_NORMAL"] = df["MARCA"].fillna("").astype(str).str.strip().str.upper()
            proformas_validas = set(
                df.loc[df["MARCA_NORMAL"].isin(marcas_objetivo), "PROFORMA"].dropna().astype(str).unique()
            )
            if not proformas_validas:
                continue
            df_filtrado = df[df["PROFORMA"].astype(str).isin(proformas_validas)].copy()
            df_filtrado = df_filtrado.sort_values(by=["PROFORMA"]).drop(columns=["MARCA_NORMAL"], errors="ignore")
            df_filtrado.to_excel(writer, sheet_name=hoja[:31], index=False)
            hubo_contenido = True

    if not hubo_contenido:
        try:
            os.remove(nombre_salida)
        except OSError:
            pass
        return None

    _colorear_por_proforma(nombre_salida, marcas_objetivo)
    logger.info(f"generar_excel_marcas: guardado en {nombre_salida}")
    return nombre_salida


def _colorear_por_proforma(ruta_excel: str, marcas_objetivo: set[str]):
    wb = load_workbook(ruta_excel)
    for hoja in wb.sheetnames:
        ws = wb[hoja]
        encabezados = {str(ws.cell(1, c).value).strip().upper(): c for c in range(1, ws.max_column + 1)}
        if "MARCA" not in encabezados or "PROFORMA" not in encabezados:
            continue
        col_marca, col_proforma = encabezados["MARCA"], encabezados["PROFORMA"]
        proformas_colores: dict[str, tuple[str, str]] = {}
        indice_color = 0
        for fila in range(2, ws.max_row + 1):
            proforma = str(ws.cell(fila, col_proforma).value).strip()
            marca = str(ws.cell(fila, col_marca).value).strip().upper()
            if proforma not in proformas_colores:
                proformas_colores[proforma] = PARES_COLORES[indice_color % len(PARES_COLORES)]
                indice_color += 1
            color_fuerte, color_suave = proformas_colores[proforma]
            color = color_fuerte if marca in marcas_objetivo else color_suave
            fill = PatternFill(fill_type="solid", start_color=color, end_color=color)
            for c in range(1, ws.max_column + 1):
                ws.cell(fila, c).fill = fill
    wb.save(ruta_excel)


def generar_excel_restringidos(ruta_historial: str, marcas_config: dict[str, set[str]]) -> str | None:
    if not ruta_historial or not os.path.exists(ruta_historial):
        return None

    fecha_hoy = datetime.now().strftime("%Y%m%d_%H%M%S")
    salida = os.path.join(REPORTES_DIR, f"restringidos_{fecha_hoy}.xlsx")

    restringidas = marcas_config["restringida_semaforo"]
    prohibidas_500_1000 = marcas_config["prohibida_500_1000"]
    excepcion_menor_500 = marcas_config["excepcion_menor_500"]
    clave = ["REQUERIMIENTO", "PROFORMA", "RUC", "CODIGO_UNICO", "FICHA_PRODUCTO"]

    xls = pd.read_excel(ruta_historial, sheet_name=None)
    hubo_contenido = False

    with pd.ExcelWriter(salida, engine="openpyxl") as writer:
        for catalogo, df in xls.items():
            if df.empty:
                continue
            df = df.drop_duplicates(subset=[c for c in clave if c in df.columns])
            df["MARCA"] = df["MARCA"].astype(str).str.strip().str.upper()
            df["COLOR_SEMAFORO"] = df["COLOR_SEMAFORO"].astype(str).str.strip().str.upper()
            df["PROCEDIMIENTO"] = df["PROCEDIMIENTO"].astype(str).str.strip().str.upper()

            mask_procedimiento = df["PROCEDIMIENTO"] == "ORDINARIA - INDIVIDUAL"
            mask_semaforo_rojo = df["COLOR_SEMAFORO"] == "ROJO"
            subtotal, marca = df["SUBTOTAL"], df["MARCA"]

            mask_semaforo = marca.isin(restringidas) & mask_semaforo_rojo
            mask_mayor_1000 = subtotal > 1000
            mask_menor_500 = (subtotal < 500) & (~marca.isin(excepcion_menor_500))
            mask_500_1000 = (subtotal >= 500) & (subtotal <= 1000) & (~marca.isin(prohibidas_500_1000))

            df_semaforo = df[mask_semaforo & mask_procedimiento]
            df_monto_minimo = df[(~mask_semaforo) & (~mask_mayor_1000) & (mask_menor_500 | mask_500_1000) & mask_procedimiento]

            if not df_semaforo.empty:
                df_semaforo.to_excel(writer, sheet_name=f"{catalogo[:24]}_semaforo", index=False)
                hubo_contenido = True
            if not df_monto_minimo.empty:
                df_monto_minimo.to_excel(writer, sheet_name=f"{catalogo[:20]}_monto_min", index=False)
                hubo_contenido = True

    if not hubo_contenido:
        try:
            os.remove(salida)
        except OSError:
            pass
        return None

    logger.info(f"generar_excel_restringidos: guardado en {salida}")
    return salida


def guardar_reporte_generado(run_id: int | None, tipo: str, ruta: str | None):
    """Registra en MySQL la ruta de un Excel ya generado en disco, para
    que el endpoint de descarga/reenvío lo pueda encontrar por run_id."""
    if not run_id or not ruta:
        return
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO perucompras_reportes_generados
                    (run_id, tipo, nombre_archivo, ruta_absoluta)
                VALUES (%s, %s, %s, %s)
                """,
                (run_id, tipo, os.path.basename(ruta), ruta),
            )
        conn.commit()
    finally:
        conn.close()


def generar_todos_los_reportes(catalogos_registros: dict[str, list[dict]], run_id: int | None = None, uid: str = ""):
    marcas_config = obtener_marcas_config()
    ruta_historial = generar_excel_historial(catalogos_registros)
    if ruta_historial:
        guardar_reporte_generado(run_id, "historial", ruta_historial)

        ruta_marcas = generar_excel_marcas(ruta_historial, marcas_config["objetivo"])
        guardar_reporte_generado(run_id, "marcas", ruta_marcas)

        ruta_restringidos = generar_excel_restringidos(ruta_historial, marcas_config)
        guardar_reporte_generado(run_id, "restringidos", ruta_restringidos)

    ruta_acumulado = generar_excel_historial_acumulado(uid)
    if ruta_acumulado:
        guardar_reporte_generado(run_id, "historial_acumulado", ruta_acumulado)